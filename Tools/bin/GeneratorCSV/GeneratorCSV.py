#-*- coding: utf-8 -*-
#coding:utf-8
import sys , getopt
import os , re , string
import time , datetime
import binascii 
import collections  
import xml.dom.minidom
import time,datetime
import socket
import csv
import traceback

#g_platform = "LINUX"
g_platform = "WIN64"
if g_platform == "WIN64":
	import ctypes
	
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.cell import get_column_letter
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter 
from openpyxl.worksheet.hyperlink import Hyperlink

from xml.dom import minidom , Node 

from xml.etree.ElementTree import ElementTree
from xml.etree.ElementTree import Element
from xml.etree.ElementTree import SubElement as SE
#sys.reload()
#sys.setdefaultencoding("utf-8") 
#sys.setdefaultencoding("cp936") 

g_yellow = ""
g_yellow_h = ""
g_green = ""
g_green_h = ""
g_red = ""
g_original = ""
g_cyan = ""
g_stdInputHandle = -10
g_stdOutputHandle = -11
g_stdErrorHandle = -12

#windows下的颜色.
FOREGROUND_BLACK = 0x0
FOREGROUND_BLUE = 0x01
FOREGROUND_GREEN = 0x02
FOREGROUND_RED = 0x04
FOREGROUND_INTENSTITY = 0x08

BACKGROUND_BLUE = 0x20
BACKGROUND_GREEN = 0x30
BACKGROUND_RED = 0x40
BACKGROUND_INTENSTITY = 0x80
			
g_isExportLua = True
g_xlsImportPath = ""
g_xlsExportCSVPath = ""
g_xlsExportServerCPPPath = ""
g_xlsExportClientCPPPath = ""

g_conditionConfigName = "_ConditionConfig"
g_commonDataObject = "Common"
g_commonDataName = "_CommonData"
g_globalFuncName = "Global"
g_xlsDeleteRecord = []
g_xlsRecords = collections.OrderedDict()
g_xlsConditionRecords = collections.OrderedDict()
g_xlsFilePath = collections.OrderedDict()
	
g_commonDatas = collections.OrderedDict()  # 这里用来记录所有的共有变量的值
g_repeatServerCommonDatas = []  # 这里用来记录所有的共有变量重复的值,防止在条件表达式中写错.
g_repeatClientCommonDatas = []  # 这里用来记录所有的共有变量重复的值,防止在条件表达式中写错.

g_serverConditions = collections.OrderedDict()  # 这里用来记录所有的条件
g_serverActions = collections.OrderedDict()
g_clientConditions = collections.OrderedDict()
g_clientActions = collections.OrderedDict()
g_serverConditionCol = 1
g_serverActionCol = 2			
g_clientConditionCol = 3
g_clientActionCol = 4
g_serverExpression = collections.OrderedDict()
g_clientExpression = collections.OrderedDict()
g_conditionObjName = "pConfig"		# 这个是当处理条件的时候.如果条件是本表内的,使用的统一名字.
g_actionNameSplit = "("				# 执行时的名字.用这个来区分
g_expressionPrefix = "Run"
g_and = "&&"	#没有或和非运算.如果想用这个.多写几个条件.
g_great = ">"
g_greatEqual = ">="
g_less = "<"
g_lessEqual = "<="
g_equal = "=="
g_notEqual = "!="

g_checkFuncPrefix = "xxCheck"
g_csvFileSuffix = ".tabcsv"
g_configPrefix = "S"
g_configSuffix = "Base"
g_loadConfigSuffix = "Load"
g_xlsNamespace = "Config"
g_conditionTag = "condition"
g_actionTag = "action"
g_pointPrefix = "p"		#指针的前缀
g_conditionSymbol = [">" , '=' , "<" , "(" , ")" , "&" , "|" , "!" , "+" , "-" , "*" , "/" , "%" , "^" , ","]

#这里为每一列的条件表达式的
g_serverCellExpression = collections.OrderedDict()
g_clientCellExpression = collections.OrderedDict()


g_serverType = "s".lower()
g_clientType = "c".lower()
g_bCheckCSTypeStrict = False	#检测CS条件的时候是否执行严格检测.即.服务器是否需要客户端数据.默认是需要的.

g_configTypePrefix = "std_unordered_map<"
g_globaleNamePrefix = "g_p"

g_int32Type = "INT32"
g_int64Type = "INT64"
g_boolType = "bool"
g_doubleType = "double"
g_stringType = "std::string"
g_dateType = "Timer::Date"
g_int32ArrayType = "std::vector<INT32>" 
g_int64ArrayType = "std::vector<INT64>"
g_boolArrayType = "std::vector<bool>"
g_doubleArrayType = "std::vector<double>"
g_stringArrayType = "std::vector<std::string>"
g_dateArrayType = "std::vector<Timer::Date>"
g_conditionType = "Condition"
g_structType = 1
g_structArrayType = 2
g_configType = 3
g_mapType = "map"
	
g_int32Func = "GetInt32"
g_int64Func = "GetInt64"
g_boolFunc = "GetBool"
g_doubleFunc = "GetDouble"
g_stringFunc = "GetString"
g_dateFunc = "GetDate"
g_int32ArrayFunc = 6
g_int64ArrayFunc = 7
g_boolArrayFunc = 8
g_doubleArrayFunc = 9
g_stringArrayFunc = 10
g_structFunc = 11
g_structArrayFunc = 12
g_dateArrayFunc = 13
g_mapFunc = 14
g_configFunc = 15

oneTab = "\t"
twoTab = oneTab + "\t"
threeTab = twoTab + "\t"
fourTab = threeTab + "\t"
fiveTab = fourTab + "\t"
sixTab = fiveTab + "\t"
sevenTab = sixTab + "\t"
eightTab = sevenTab + "\t"

g_cellID = 0  	# 第一列为ID列
g_rowComent = 0 # 注释行
g_rowType = 1	# 类型行
g_rowName = 2 	# 为类型名字行
g_rowCheck = 3 	# 第三行为检测数据正确性行,以及动态填写数据行.
g_rowCS = 4   	# 第四行为是否是服务器还是客户端代码
g_rowDataStart=5# 从这行开始就是数据了.

g_conditionRowName = 0 	# 条件列的   条件名字
g_conditionRowCheck = 1 		# 条件列的   保持数据兼容.填写NULL,
g_conditionRowCS = 2 		# 条件列的   服务器还是客户端代码.
g_conditionDataStart = 3 	# 条件列的   数据开始行.

class Object(object):
	"""docstring for Object"""
	def __init__(self, arg):
		super(Object, self).__init__()
		self.name = None # 是否有对象,或者静态函数没有对象
		self.namespace = []

class Action(object):
	"""docstring for Action"""
	def __init__(self, arg):
		super(Action, self).__init__()
		self.actionName = Object(None)
		self.args = []
		
class Condition(object):
	"""docstring for Condition"""
	def __init__(self, arg):
		super(Condition, self).__init__()
		self.conditionName = Object(None)
		self.args = []
		
class ExpressionCondition(object):
	"""docstring for ExpressionCondition"""
	def __init__(self, arg):
		super(ExpressionCondition, self).__init__()
		self.text = ""
		self.objs = []
		
class ExpressionAction(object):
	"""docstring for ExpressionAction"""
	def __init__(self, arg):
		super(ExpressionAction, self).__init__()
		self.name = ""
		self.args = []
		self.objs = []
		
class Expression(object):
	"""docstring for Expression"""
	def __init__(self, arg):
		super(Expression, self).__init__()
		self.condition = ExpressionCondition(None)
		self.actions = []		

class CommonData(object):
	"""docstring for CommonData"""
	def __init__(self, arg):
		super(CommonData, self).__init__()
		self.value = ""
		self.keyType = ""
		self.valueType = ""
		self.objects = []

def start(): 
	LogOutInfo("start generate csv.\n")  
	DeleteExportPathFiles() 
	CreateExportPathFiles()
	GenerateCSVFromXLS()
	#CheckXLS()
	LogOutInfo("generate CSV finished.\n") 		
	HandleConditionConfig()
	HandleCommondataConfig()
	HandleSheetCondition()
	CheckRepeatCommonData()
	HandleSheetCellCondition()
			
	LogOutInfo("start generate CPP.\n")   	
	GenerateCPP()
	LogOutInfo("generate CPP finished.\n") 
	
def CheckXLS():
	root = g_xlsImportPath
	files = []
	Search(root , '.xlsx' , files)
	for filepath in files:
		HyperLinkXLS(filepath)

def HyperLinkXLS(filepath):
	dirout = g_xlsExportCSVPath
	dirout = dirout + os.sep  #新路径名称

	filename = os.path.basename(filepath) #获取文件名
	filename = os.path.splitext(filename.replace(' ', '_'))[0]

	bWrite = False
	xlsx_file = load_workbook(filepath , False , False , False , False , False )
	writeWB = ExcelWriter(xlsx_file)
	for sheet in xlsx_file.get_sheet_names():		
		cur_real_row_index = -1
		sheet_ranges = xlsx_file[sheet]
		for row in sheet_ranges.rows:
			cur_real_row_index += 1

			cur_real_cell_index = -1
			for cell in row:
				if cur_real_row_index != g_rowComent:
					break
				cur_real_cell_index = cur_real_cell_index + 1
				
				hyperLink = ""				
				item = sheet_ranges.rows[g_rowType][cur_real_cell_index].value
				#LogOutDebug("filename:" , filename , " cur_cell_index:" , cur_cell_index , " len:" , len(g_xlsRecords[filename][g_rowType]))
				#LogOutDebug("item:" , item )
				colType = GetType(item)
				if colType == g_configType:
					hyperLink = GetDicKeyByUpper(g_xlsRecords , item) # 找绝对路径.
				elif colType == g_structType or colType == g_structArrayType:
					item = RemoveSpecialWord(item)
					childItems = item.split(',')
					for indexChild , childItem in enumerate(childItems):
						childType = GetType(childItem)
						if childType == g_configType:
							hyperLink = GetDicKeyByUpper(g_xlsRecords , childItem) # 找绝对路径.
							break
				elif colType.find(g_mapType) >= 0:
					mapType , keyName , valueName = GetMapType(item , False)
					if GetType(keyName) == g_configType:
						hyperLink = keyName
					elif GetType(valueName) == g_configType:
						hyperLink = valueName

				if len(hyperLink) > 0:
					bWrite = True
					LogOutDebug("hyperLink:" , hyperLink)
					sheet_ranges.rows[g_rowComent][cur_real_cell_index].hyperlink = g_xlsFilePath[hyperLink]
					#sheet_ranges.rows[g_rowComent][cur_real_cell_index].style.fill.fill_type = 'solid'
					#sheet_ranges.rows[g_rowComent][cur_real_cell_index].style.fill.start_color = Color("FF0000")
					#sheet_ranges.rows[g_rowComent][cur_real_cell_index].style.fill.end_color = Color("00FF00")
	if bWrite:
		writeWB.save(filepath)
	#LogOutDebug("g_xlsConditionRecords:" , len(g_xlsConditionRecords[filename]))

		
def GenerateCSVFromXLS():
	root = g_xlsImportPath
	files = []
	Search(root , '.xlsx' , files)
	for result in files:
		LogOutInfo("load file: " , result);
		Xlsx2CSV(result)

	CheckRecords()
	GenerateCSV()

def Xlsx2CSV(filepath):
	dirout = g_xlsExportCSVPath
	dirout = dirout + os.sep  #新路径名称
	try:
		# 一个表中的所有sheet输出到一个csv文件中,所以要保证格式一致.文件名用xlsx文件名
		filename = os.path.basename(filepath) #获取文件名
		filename = os.path.splitext(filename.replace(' ', '_'))[0]
		#LogOutDebug("path:" , filepath)
		g_xlsFilePath[filename] = os.path.abspath(filepath)
		if filename.find('#') >= 0:
			filename = filename.replace('#' , '')
			LogOutInfo("skip file: " , filepath )
			g_xlsDeleteRecord.append(filename)
		
		id_list = []			# 用于重复的ID去除
		delete_col_list = []	# 不读取需要删除的列
		cur_sheet_index = 0
		xlsx_file_reader = load_workbook(filepath , True , False , False , False , True )
		g_xlsRecords[filename] = collections.OrderedDict()
		conditionContainer = collections.OrderedDict() #这个专门处理条件表达式的.
		for sheet in xlsx_file_reader.get_sheet_names():			
			cur_rows_index = 0
			sheet_ranges = xlsx_file_reader[sheet]
			for row in sheet_ranges.rows:
				if cur_sheet_index >= 1 and cur_rows_index < g_rowDataStart + 1: #当第二个sheet的时候前4行是不读取的.
					cur_rows_index = cur_rows_index + 1
					continue
				
				row_container = [] 
				cur_cell_index = 0
				for cell in row:		
					Str = ""
					cellValue = cell.value
					if type(cellValue) == type(None):
						if cur_cell_index == g_cellID:
							break		
						Str = ""	
#						LogOutError("error parase filepath" , filepath , "  cur_sheet " , sheet , "  cur_rows_index " , cur_rows_index ,"  cur_cell_index " , cur_cell_index , "  type(cell.value) " , type(cell.value))
					else:
						Str = cellValue
						if type(cellValue) != str:
							Str = str(cellValue)
						else:
							Str = Str.encode('gbk').decode('gbk').encode('utf-8').decode('utf-8')

					if cur_rows_index == g_rowComent and len(Str) > 0 and Str.lower()[0] == '#':  # 跳过这一列
						#LogOutDebug("delete_col_list" , delete_col_list)
						delete_col_list.append(cur_cell_index)
					if cur_cell_index in delete_col_list:
						cur_cell_index = cur_cell_index + 1
						#LogOutDebug("cur_cell_index continue" , cur_cell_index)
						continue

					if cur_cell_index == g_cellID and len(Str) > 0 and Str.lower()[0] == '#':  # 跳过这一行
						#LogOutDebug("cur_cell_index break " , cur_cell_index)
						cur_cell_index = cur_cell_index + 1
						break

					if len(Str) >= 0:			#有些策划没填的数据为空.也需要记录.
						if cur_rows_index < 4:
							Str = ''.join([x for x in Str if x != " "]) 
						if cur_cell_index == g_cellID:		# 第一列是ID.需要判定是否有重复的ID
							if Str in id_list:
								LogOutDebug("repeat id \'" , Str , "\' in \'" , filepath , " \' file")
								#LogOutError("repeat id \'" , Str , "\' in \'" , filepath , " \' file")
							id_list.append(Str)
						if cur_rows_index == g_rowType and Str.lower() == g_conditionType.lower():
							conditionContainer[cur_cell_index] = []  #这个用来记录这一列的条件表达式.
						else:
							if cur_cell_index in conditionContainer:
								if len(Str) > 0:  # 记录条件的时候因为不是一一对应的,所以当为空的时候不需要记录.
									conditionContainer[cur_cell_index].append(Str)
									RemovListNewLine(conditionContainer[cur_cell_index])
									#LogOutDebug("cur_cell_index:" , cur_cell_index , " conditionContainer:" , len(conditionContainer[cur_cell_index]) , " Str:" , str(Str))
							else:
								row_container.append(Str)

					cur_cell_index = cur_cell_index + 1

				if len(row_container) >= 1:	
					RemovListNewLine(row_container)
					g_xlsRecords[filename][cur_rows_index] = row_container
					cur_rows_index = cur_rows_index + 1   #这里是为了去除一些空行.
					#LogOutDebug("cell.row_container:" , row_container)
			cur_sheet_index = cur_sheet_index + 1	
		g_xlsConditionRecords[filename] = conditionContainer
		#LogOutDebug("g_xlsConditionRecords:" , len(g_xlsConditionRecords[filename]))

	except Exception as e:
		LogOutError(e)
		
def CheckRecords():
	for sheet , item in g_xlsRecords.items(): 	#读取sheet
		#LogOutDebug("sheet" , sheet)
		for row , rowItem in item.items():		#读取每一行
			if row == g_rowName:				#命名重名检测,
				sameName = []
				for col , colItem in enumerate(rowItem):	#读取每一列
					colItemName = colItem
					item_type = GetType(g_xlsRecords[sheet][g_rowType][col])
										
					if colItem[0] == '[' or colItem[0].isdigit():
						LogOutError("sheet=" , sheet , " :row=" , row , " :col=" , col , " item=" , colItem , "use error name .")
					if item_type == g_structArrayType:
						if colItem.find('[') <= 0 or colItem.find(']') <= 0:
							LogOutError("sheet=" , sheet , " :row=" , row , " :col=" , col , " item=" , colItem , "error  .no \'[\' or \']\'.")
					
					if item_type == g_structArrayType or item_type == g_structType:
						npos = colItem.find('[')
						colItemName = colItem[0 : npos]
					
					if colItemName in sameName:
						LogOutError("sheet=" , sheet , " :row=" , row , " :col=" , col , " item=" , colItemName , " use one same name .")
					else:
						sameName.append(colItemName)						
			elif row == g_rowType:		#类型
				for col , colItem in enumerate(rowItem):	#读取每一列
					item_type = GetType(colItem)					
				
			elif row == g_rowCS:	
				for col , colItem in enumerate(rowItem):	#读取每一列
					if col == g_cellID:
						g_xlsRecords[sheet][row][col] = g_serverType + g_clientType
					CheckCSType(colItem)	

			elif row == g_rowComent:	
				pass	
			elif row == g_rowCheck:	
				pass	
			else:
				#LogOutDebug("g_xlsRecords,[row = " + str(row) + ":" , g_xlsRecords[sheet][row])
				for col , colItem in enumerate(rowItem):	#读取每一列
					item_type = GetType(g_xlsRecords[sheet][g_rowType][col])	
					#LogOutDebug("CheckDataType,row=" + str(row) + ":item_type=" , item_type , "sheet" , sheet , "col=" , col , " colItem:" , colItem)
					CheckDataType(item_type , sheet , row , col , colItem , -1)
					#LogOutDebug("CheckDataType,row=" + str(row) + ":item_type=" , item_type)
					
	for sheet , item in g_xlsConditionRecords.items():
		for col , colItem in item.items():		# 读取每一列
			sameName = []						# 命名重名检测,
			for row , rowItem in enumerate(colItem):		# 读取每一行
				if row == 0:	# 第一行是名字
					sameName.append(rowItem)
				elif rowItem in sameName:
					LogOutError("sheet=" , sheet  , " :row=" , row , " :col=" , col , " item=" , rowItem , " use one same name .")

			for row , rowItem in enumerate(colItem):	#读取每一列
				CheckDataType(g_conditionType , sheet , row , col , rowItem , -1)
				#LogOutDebug("g_xlsConditionRecords:" , g_xlsConditionRecords[sheet][col][row])
				
def HandleConditionFromItem(conditions , colItem):
	args = []
	colItem = RemoveSpecialWord(colItem)
	condition = colItem.split('(')		# 这里至少是2个.
	if len(condition) > 1:
		args = condition[1]
	
	# 处理名字.包括命名空间
	names = condition[0].split('::')
	name = names[len(names) - 1]			# 这个肯定是名字
	if name not in conditions:
		conditions[name] = Condition(None)
		conditions[name].conditionName.name = name
	for index , obj in enumerate(names):
		if index != len(names) - 1:
			conditions[name].conditionName.namespace.append(obj)
		
	#处理条件的参数
	if len(args) > 0:
		args = args.replace(")" , "").split(',')
		#LogOutDebug("args1:" , args)
	if len(args) <= 0 or len(args[0]) <= 0:
		args = []		
		#LogOutDebug("args2:" , args)
	for index , arg in enumerate(args):
		argNames = arg.split('::')
		argObject = Object(None)
		if len(argNames) > 0:
			argName = argNames[len(argNames) - 1]			# 这个肯定是名字
		argObject.name = argName
		for argIndex , argName in enumerate(argNames):
			if argIndex != len(argNames) - 1:
				argObject.namespace.append(argName)
		conditions[name].args.append(argObject)	
		
def HandleActionFromItem(actions , colItem):	
	args = []
	colItem = RemoveSpecialWord(colItem)
	action = colItem.split('(')		# 这里至少是2个.
	if len(action) > 1:
		args = action[1]
	
	# 处理名字.包括命名空间
	names = action[0].split('.')
	if len(names) != 1 and len(names) != 2:
		LogOutError("HandleActionFromItem name=" , action[0] , " has multi object or no object.")
		
	objName = None
	name = None
	namespaces = None
	if len(names) == 1:
		namespaces = action[0].split('::')	
		if len(namespaces) == 1:
			LogOutError("HandleActionFromItem name=" , colItem , " error. namespace or obj error.")
		else:
			name = namespaces[len(namespaces) - 1]			# 这个肯定是名字
	else:
		name = names[1]			# 这个肯定是名字
		namespaces = names[0].split('::')
		objName = namespaces[len(namespaces) - 1]			# 这个肯定是对象名字	
	
	namespaces = names[0].split('::')
	if name not in actions:
		actions[name] = Action(None)
		actions[name].actionName.name = objName
	for index , obj in enumerate(namespaces):
		if index != len(namespaces) - 1:	  # 对于action而言.括号前的倒数第1个参数代表对象或者是函数名字.CUtil::Player.Mail(CUtil.League , CUtil.Player) or CUtil::MinValue
			actions[name].actionName.namespace.append(obj)
	
	#处理条件的参数
	if len(args) > 0:
		args = args.replace(")" , "").split(',')
		#LogOutDebug("args1:" , args)
	if len(args) <= 0 or len(args[0]) <= 0:
		args = []		
		#LogOutDebug("args2:" , args)
	for index , arg in enumerate(args):
		argNames = arg.split('::')
		argObject = Object(None)
		if len(argNames) > 0:
			argName = argNames[len(argNames) - 1]			# 这个肯定是名字
		argObject.name = argName
		for argIndex , argName in enumerate(argNames):
			if argIndex != len(argNames) - 1:
				argObject.namespace.append(argName)
		actions[name].args.append(argObject)	
		
def HandleConditionConfig():
	rowItems = g_xlsRecords[g_conditionConfigName]
	if len(rowItems) != 0:
		for row , rowItem in rowItems.items():				# 读取每一行
			for col , colItem in enumerate(rowItem):		# 读取每一列
				if row >= g_rowDataStart: 					# 从这行开始是数据	
					#LogOutInfo("row:" , row , " col:" , col , " colItem:" , colItem)	
					if len(colItem) == 0:
						continue			
					if col == g_serverConditionCol:
						HandleConditionFromItem(g_serverConditions , colItem)						
					if col == g_serverActionCol:
						HandleActionFromItem(g_serverActions , colItem)						
					if col == g_clientConditionCol:
						HandleConditionFromItem(g_clientConditions , colItem)						
					if col == g_clientActionCol:
						HandleActionFromItem(g_clientActions , colItem)	
		#LogOutDebug("g_serverConditions:" , g_serverConditions)			
		#LogOutDebug("g_serverActions:" , g_serverActions)
		#LogOutDebug("g_clientConditions:" , g_serverConditions)
		#LogOutDebug("g_clientActions:" , g_clientActions)
	else:
		LogOutError("HandleConditionConfig error." , g_conditionConfigName , " no data")

def HandleCommondataConfig():
	rowItems = g_xlsRecords[g_commonDataName]
	if len(rowItems) != 0:
		for row , rowItem in rowItems.items():				# 读取每一行
			for col , colItem in enumerate(rowItem):		# 读取每一列
				if row >= g_rowDataStart: 					# 从这行开始是数据	
					if len(colItem) == 0:
						continue
					if col <= g_cellID:
						continue
					mapType , keyType , valueType = GetMapType(g_xlsRecords[g_commonDataName][g_rowType][col])
					#LogOutDebug("row:" , row , " col:" , col , " colItem:" , colItem , " mapType:" , mapType , " keyType:", keyType ," valueType:" , valueType)	

					datas = colItem.split('=')
					commonKey = datas[0]
					if commonKey not in g_commonDatas:
						g_commonDatas[commonKey] = CommonData(None)
						g_commonDatas[commonKey].value = datas[1]
						g_commonDatas[commonKey].keyType = keyType
						g_commonDatas[commonKey].valueType = valueType
						g_commonDatas[commonKey].objects = []
						g_commonDatas[commonKey].objects.append(g_commonDataObject)
						#g_commonDatas[datas[0]].objects.append(g_xlsRecords[g_commonDataName][g_rowName][col])
					else:
						LogOutError("row:" , row , " col:" , col , " colItem:" , colItem , " mapType:" , mapType , " keyType:", keyType ," valueType:" , valueType , " error. the sameName" , datas[0])
					
	else:
		LogOutError("HandleConditionConfig error." , g_conditionConfigName , " no data")
	#LogOutDebug("g_commonDatas:" , g_commonDatas)

def CheckRepeatCommonData():
	for conmonIndex , commondata in g_commonDatas.items():
		for serverIndex , serverCondition in g_serverConditions.items():
			#LogOutDebug(conmonIndex , " commondata=" , commondata , " serverCondition=" , serverCondition)
			if conmonIndex.upper() == serverIndex.upper():
				g_repeatServerCommonDatas.append(conmonIndex)
				LogOutDebug("the same commondata warning." , conmonIndex , " commondata=" , commondata , " serverCondition=" , serverCondition)
		for clientIndex , clientCondition in g_clientConditions.items():
			#LogOutDebug(conmonIndex , " commondata=" , commondata , " clientCondition=" , clientCondition)
			if conmonIndex.upper() == clientIndex.upper():
				g_repeatClientCommonDatas.append(conmonIndex)
				LogOutError("the same commondata warning." , conmonIndex , " commondata=" , commondata , " clientCondition=" , clientCondition)

def HandleSheetCondition():	
	for sheet , item in g_xlsConditionRecords.items():
		g_serverExpression[sheet] = collections.OrderedDict()			# 为每一个sheet创建一系列的条件表达式
		g_clientExpression[sheet] = collections.OrderedDict()			# 为每一个sheet创建一系列的条件表达式
		for col , colItem in item.items():		# 读取每一列
			bConditionStart = False
			serverExpression = None
			clientExpression = None
			cellName = ""
			#LogOutDebug("sheet:" , sheet)
			csType = g_xlsConditionRecords[sheet][col][g_conditionRowCS]
			bServer = CheckCSType(csType , True , True)
			bClient = CheckCSType(csType , False , True)
			for row , rowItem in enumerate(colItem):		# 读取每一行
				item = RemoveSpecialWord(rowItem)
				#LogOutDebug("condition row:" , str(row) )
				if row == g_conditionRowName:
					cellName = g_expressionPrefix + rowItem
					if bServer:
						#LogOutDebug("cellName:" , cellName , "g_serverExpression[sheet]:" , g_serverExpression[sheet])
						g_serverExpression[sheet][cellName] = []	# 这里用来存每一个表达式
					if bClient:
						g_clientExpression[sheet][cellName] = []	# 这里用来存每一个表达式
				if row >= g_conditionDataStart:
					tagitem = item.lower()[0:item.lower().find(':')]
					itemContent = item.lower()[item.lower().find(':') + 1:]
					#LogOutDebug("tagitem=" , tagitem.lower() , " itemContent " , itemContent.lower())
					if tagitem.lower() == g_conditionTag.lower():
						if serverExpression != None:
							g_serverExpression[sheet][cellName].append(serverExpression)
						if clientExpression != None:
							g_clientExpression[sheet][cellName].append(clientExpression)
						#LogOutDebug("condition csType:" , csType , " IsServer=" , IsServerCSType(csType) , " itemContent " , itemContent)
						if bServer:
							serverExpression = Expression(None)
							serverExpression = HandleConditionExpression(True , sheet , itemContent , serverExpression)
							serverExpression.actions = []
						if bClient:
							clientExpression = Expression(None)
							clientExpression = HandleConditionExpression(False , sheet , itemContent , clientExpression)
							clientExpression.actions = []
							#LogOutDebug("1.serverExpression:" , serverExpression.condition)
							#LogOutDebug("1.clientExpression:" , clientExpression.condition)
					elif tagitem.lower() == g_actionTag.lower():
						if bServer:
							action = HandleExpressionAction(True , itemContent)
							serverExpression.actions.append(action)
							#LogOutDebug("action=" , action)
						if bClient:
							action = HandleExpressionAction(False , itemContent)
							clientExpression.actions.append(action) 
							#LogOutDebug("action=" , action)
					else:
						LogOutError("sheet=" , sheet , " :row=" , row , " :col=" , col , " item=" , item , " condition tag error.")				
			
			if cellName != "":
				if bServer and serverExpression != None:
					g_serverExpression[sheet][cellName].append(serverExpression)	
					#LogOutDebug("serverExpression:" , serverExpression.condition)
				if bClient and clientExpression != None:
					g_clientExpression[sheet][cellName].append(clientExpression)	
					#LogOutDebug("clientExpression:" ,clientExpression.condition)

def HandleSheetCellCondition():		
	for sheet , item in g_xlsRecords.items():	
		datas = item[g_rowName]	
		g_serverCellExpression[sheet] = collections.OrderedDict()			# 为每一个sheet创建一系列的条件表达式
		g_clientCellExpression[sheet] = collections.OrderedDict()			# 为每一个sheet创建一系列的条件表达式
	
		#LogOutDebug("sheet:" , sheet)
		for index , col in enumerate(item[g_rowType]):		# 读取数据
		
			csType = item[g_rowCS][index]
			bServer = CheckCSType(csType , True , True)
			bClient = CheckCSType(csType , False , True)
			#LogOutDebug("condition index:" , str(index) )
			
			cellName = ""
			item_type = GetType(col)
			if item_type == g_structType or item_type == g_structArrayType:
				npos = datas[index].find('[')
				cellName = datas[index][0 : npos]		
			else:
				cellName = datas[index]

			serverExpression = None
			clientExpression = None
			if bServer:
				#LogOutDebug("cellName:" , cellName , "g_serverExpression[sheet]:" , g_serverExpression[sheet])
				g_serverCellExpression[sheet][cellName] = []	# 这里用来存每一个表达式
			if bClient:
				g_clientCellExpression[sheet][cellName] = []	# 这里用来存每一个表达式
			
			expressionContent = item[g_rowCheck][index]			# 获取这一列的表达式内容
			if expressionContent.upper() == "NULL" or expressionContent.upper() == "":
				continue
			conditionContent = expressionContent.split("condition:".lower())
			
			#LogOutDebug("conditionContent:" , conditionContent)
			for i , condition in enumerate(conditionContent):
				if condition == "":
					continue
				actionContent = condition.split("action:".lower())
				expressionCondition = actionContent[0]				# 这个一定是条件
				if bServer:
					serverExpression = Expression(None)
					serverExpression = HandleConditionExpression(True , sheet , expressionCondition , serverExpression , True)
					serverExpression.actions = []
					#LogOutDebug("1.serverExpression:" , serverExpression.condition)
				if bClient:
					clientExpression = Expression(None)
					clientExpression = HandleConditionExpression(False , sheet , expressionCondition , clientExpression , True)
					clientExpression.actions = []
					#LogOutDebug("2.clientExpression:" , clientExpression.condition)
					
				# 处理Action
				for j , actionCall in enumerate(actionContent):
					if j == 0:
						continue
					if bServer:
						action = HandleExpressionAction(True , actionCall)
						serverExpression.actions.append(action)
						#LogOutDebug("S action=" , action)
					if bClient:
						action = HandleExpressionAction(False , actionCall)
						clientExpression.actions.append(action) 
						#LogOutDebug("C action=" , action)
							
				if cellName != "":
					if bServer and serverExpression != None:
						g_serverCellExpression[sheet][cellName].append(serverExpression)	
						#LogOutDebug("cellName=" , cellName , ":serverExpression=" , g_serverCellExpression[sheet][cellName])
					if bClient and clientExpression != None:
						g_clientCellExpression[sheet][cellName].append(clientExpression)	
						#LogOutDebug("cellName=" , cellName , ":clientExpression=" , g_serverCellExpression[sheet][cellName])

def HandleConditionExpression(bServer , sheet , itemContent , expression , bCell = False):	
	result = ""
	start = 0
	end = 0
	keyWord = ""
	symbol = True
	objs = []
	for i , item in enumerate(list(itemContent)):
		if item in g_conditionSymbol:
			end = i
			if end != start: # 代表是一个符号					
				keyWord = itemContent[start:end]
				#LogOutDebug("keyWord:" , keyWord , "result:" , result)
				value , symbol = ParaseConditionKeyWord(bServer , keyWord.upper() , sheet , item , objs, bCell)
				result = result + value
			if symbol == True:
				result += item
			else:
				symbol = True
			start = i + 1			
	keyWord = itemContent[start:]
	#LogOutDebug("keyWord" , keyWord)
	if len(keyWord) > 0:# 最后一个肯定不是符号.所以必然会有一个条件.
		value , symbol = ParaseConditionKeyWord(bServer , keyWord.upper() , sheet , "" , [] , bCell)
		result = result + value
	expression.condition.text = result
	expression.condition.objs = objs
	#LogOutDebug("bServer:" + str(bServer) + ":result." , result)

	return expression	
	
def ParaseConditionKeyWord(bServer , keyWord , sheet , symbol , objs , bCell = False):
	keys = keyWord.split(".")
	last = keys[len(keys) - 1].upper()
	result = ""	
	
	dicCommondata = collections.OrderedDict()
	dicConditions = collections.OrderedDict()
	dicOrigCondition = collections.OrderedDict()
	dicRepeatCommondata = []
	
	MakeDicKeyUpper(g_commonDatas , dicCommondata)
	if bServer:
		dicRepeatCommondata = g_repeatServerCommonDatas
		dicOrigCondition = g_serverConditions
		MakeDicKeyUpper(g_serverConditions , dicConditions)
	else:
		dicRepeatCommondata = g_repeatClientCommonDatas
		dicOrigCondition = g_clientConditions
		MakeDicKeyUpper(g_clientConditions , dicConditions)
		
	isInCommonData = False
	if last in dicCommondata:
		isInCommonData = True
		commondata = dicCommondata[last]
		for index , key in enumerate(keys):
			if key in commondata.objects:
				if commondata.valueType != g_dateType:
					return commondata.value , True
				else:
					return g_dateType + "(\"" + commondata.value + "\")" , True

	if last not in dicRepeatCommondata:
		if isInCommonData:
			commondata = dicCommondata[last]
			if commondata.valueType != g_dateType:
				return commondata.value , True
			else:
				return g_dateType + "(\"" + commondata.value + "\")" , True

		if last in dicConditions:
			condition = dicConditions[last]
			return GetConditionByCondition(bServer , condition , GetDicKeyByUpper(dicOrigCondition , last) , symbol , objs)
	else:
		condition = dicConditions[last]
		isInCondition = False
		for index , key in enumerate(keys):
			if key in condition.objs:
				isInCondition = True
		if isInCondition:
			return GetConditionByCondition(bServer , condition , GetDicKeyByUpper(dicOrigCondition , last) , symbol , objs)

	for index , rowName in enumerate(g_xlsRecords[sheet][g_rowName]):
		#LogOutDebug("rowName:", rowName.upper() , " last:" , last , ":cell=" , bCell) 
		if rowName.upper() == last:
			if not bCell:
				return "(" + "Get" + sheet + "(id)->" + rowName + ")" , True
			else:
				return "conf." + rowName + "" , True
				
	#LogOutError("ParaseConditionKeyWord ERR." , keyWord , "not in commondata or condition or cur excel")

def GetConditionByCondition(bServer , condition , last , symbol , objs):
	needWriteSymbol = True
	result = ""
	if bServer:
		result = result + "CUtil::Condition<"
	else:
		result = result + "CUtil::Condition<"
	for index , obj in enumerate(condition.conditionName.namespace):
		result = result + obj
		result = result + "::"
	result = result + last
	result = result + ">()"
	
	if not (symbol == '(' and len(condition.args) <= 0): # 如果下一个是括号.但是又没有对象.则不添加这个括号.
		result = result + "("
	for index , arg in enumerate(condition.args):
		result += g_pointPrefix
		result += arg.name
		#LogOutDebug("condition.args:" , condition.args)
		if len(condition.args) - 1 != index:
			result = result + ","
		else:
			if symbol == '(':
				result = result + ","
		bInArgs = False
		for argIndex , obj in enumerate(objs):
			if arg.name == obj.name:
				bInArgs = True
		if not bInArgs:
			objs.append(arg)
			
	if symbol != '(':
		result = result + ")"
	else:
		if len(condition.args) > 0:
			needWriteSymbol = False
		
	return result , needWriteSymbol


def HandleExpressionAction(bServer , itemContent):	
	action = ExpressionAction(None)
	actions = itemContent.split(g_actionNameSplit)
	if len(actions) > 0:
		args = []
		action.name = CheckActionValue(bServer , actions[0] , action.objs , action.args)
		actions[1] = RemoveSpecialWord(actions[1])
		actions[1] = actions[1].replace(")" , "").replace("(" , "")
		action.args += actions[1].split(',')
			 
	return action

def CheckActionValue(bServer ,  value , objs , args):	
	tmp = collections.OrderedDict()
	if bServer:
		MakeDicKeyUpper(g_serverActions , tmp)
	else:
		MakeDicKeyUpper(g_clientActions , tmp)
	
	if value.upper() in tmp:
		result = ""
		name = tmp[value.upper()].actionName.name
		if name != None and len(name) > 0:
			result += g_pointPrefix
			result += name
			result += "->"
				
			bIn = False
			for objIndex , obj in enumerate(objs):
				if tmp[value.upper()].actionName.name == obj.name:
					bIn = True
			if not bIn:
				objs.append(tmp[value.upper()].actionName)
			
			for index , arg in enumerate(tmp[value.upper()].args):
				argValue = g_pointPrefix
				argValue += arg.name
				args.append(argValue)
				
				bIn = False
				for objIndex , obj in enumerate(objs):
					if arg.name == obj.name:
						bIn = True
				if not bIn:
					objs.append(arg)
		elif name == None:
			for index , namespace in enumerate(tmp[value.upper()].actionName.namespace):
				result += namespace				
				result += "::"
			
		if bServer:
			result += GetDicKeyByUpper(g_serverActions , value.upper())
		else:
			result += GetDicKeyByUpper(g_clientActions , value.upper())
		return result
	else:
		LogOutError("CheckActionValue error." , value , " tmp:" , tmp)
	

def GenerateCSV():
	
	dirout = g_xlsExportCSVPath
	dirout = dirout + os.sep  #新路径名称
	try:
		for sheet , item in g_xlsRecords.items(): 	#读取sheet
			filename = sheet
			csv_filename = ('{xlsx}' + g_csvFileSuffix).format(xlsx=filename)		
			dirfileout  = dirout + csv_filename
			#LogOutDebug("dirfileout" , dirfileout )
			csv_file = open(dirfileout , 'w' , newline='' , encoding='utf_8_sig')
			
			#QUOTE_MINIMAL QUOTE_NONE所有的都要加引号.
			csv_file_writer = csv.writer(csv_file , delimiter='	', quotechar='"', quoting=csv.QUOTE_ALL)

			for row , rowItem in item.items():	#读取每一行
				csv_file_writer.writerow(rowItem)

			csv_file.close()
	except Exception as e:
		LogOutError(e)

def GenerateCPP(): 
	GenerateCPPFiles(True)
	GenerateCPPFiles(False)
	
def ExportLua(): 
	ExportLuaFiles(True)
	ExportLuaFiles(False)
	
def ExportLuaFiles(bServer): 
	ExportClassToLua(bServer)
		
def GenerateCPPFiles(bServer): 
	GenerateConfigManagerHeader(bServer)
	GenerateConfigDeclareHeader(bServer)
	for sheet , item in g_xlsRecords.items():
		GenerateConfigLoadHeader(bServer , sheet , item[g_rowType] , item[g_rowName] , item[g_rowComent] , item[g_rowCheck] , item[g_rowCS])
		GenerateConfigLoadCpp(bServer , sheet , item[g_rowType] , item[g_rowName] , item[g_rowComent] , item[g_rowCheck] , item[g_rowCS])
		GenerateConfigBaseHeader(bServer , sheet , item[g_rowType] , item[g_rowName] , item[g_rowComent] , item[g_rowCS])
		GenerateConfigBaseCpp(bServer , sheet , item[g_rowType] , item[g_rowName] , item[g_rowComent] , item[g_rowCS])
		GenerateConfigHeader(bServer , sheet , item[g_rowType] , item[g_rowName] , item[g_rowComent] , item[g_rowCS])
		GenerateConfigCpp(bServer , sheet , item[g_rowType] , item[g_rowName] , item[g_rowComent] , item[g_rowCS])
	GenerateConfigManagerCPP(bServer)
	
def GenerateStructData(bServer , fileWrite , types , datas , comments , css , bLoad):	
	for index , item in enumerate(types):
		#LogOutDebug("bServer:" , bServer , " data:" , datas[index] , "css:" , css[index])
		if not CheckCSType(css[index] , bServer):
			continue
		item_type = GetType(item)
		if item_type == g_structType:
			npos = datas[index].find('[')
			structName = datas[index][0 : npos]
			structDatas = datas[index][npos + 1 : len(datas[index]) - 1].split(',')
			#LogOutDebug("name=" , structName , ":structDatas " , structDatas , " size=" , len(structDatas))
			fileWrite.write("\n" + twoTab + "//" + comments[index] + "\n") 
			fileWrite.write(twoTab + "struct " + g_configPrefix + structName + "\n");
			fileWrite.write(twoTab + "{\n");
			childItems = item.split(',')
			if len(childItems) != len(structDatas):
				LogOutError("parase struct error.maybe no detail name. structName=" , structName , ":childItems " , childItems , "size=" , len(childItems) , ":structDatas " , structDatas , " size=" , len(structDatas))
			for indexChild , childItem in enumerate(childItems):
				#LogOutDebug("name=" , structName , ":structDatas=" , structDatas , ":childItem=" , childItem)

				childType = GetType(childItem)
				if childType == g_configType:
					parentName = GetDicKeyByUpper(g_xlsRecords , childItem)
					childType , keyType , valueType  = GetConfigType(parentName)
					if bLoad:
						childType = GetType(keyType)

				fileWrite.write(threeTab + childType + GetTypeTab(childItem) + structDatas[indexChild] + ";\n")
			fileWrite.write(twoTab + "}" + structName + ";\n");
		elif item_type == g_structArrayType:
			npos = datas[index].find('[')
			structName = datas[index][0 : npos]
			structDatas = datas[index][npos + 1 : len(datas[index]) - 1].split(',')
			fileWrite.write("\n" + twoTab + "//" + comments[index] + "\n") 
			fileWrite.write(twoTab + "struct " + g_configPrefix + structName + "\n");
			fileWrite.write(twoTab + "{\n");
			item = item.replace('[' , '').replace(']' , '')
			childItems = item.split(',')
			if len(childItems) != len(structDatas):
				LogOutError("parase struct error.invalid size . name=" , structName , ":childItems " , childItems , "size=" , len(childItems) , ":structDatas " , structDatas , " size=" , len(structDatas))
			for indexChild , childItem in enumerate(childItems):
				childType = GetType(childItem)
				if childType == g_configType:
					parentName = GetDicKeyByUpper(g_xlsRecords , childItem)
					childType , keyType , valueType  = GetConfigType(parentName)
					if bLoad:
						childType = GetType(keyType)

				fileWrite.write(threeTab + childType + GetTypeTab(childItem) + structDatas[indexChild] + ";\n")
			fileWrite.write(twoTab + "}" + ";\n");
			fileWrite.write(twoTab + "std::vector<" + g_configPrefix + structName + ">" + twoTab + "vec" + structName + ";\n");
		else:
			if item_type == g_configType:
				item_type , keyType , valueType = GetConfigType(item)
				if bLoad:
					item_type = GetType(keyType)
				
			fileWrite.write(twoTab + item_type + GetTypeTab(item) + oneTab + datas[index] + ";"  + oneTab)
			fileWrite.write("//" + comments[index] + "\n") 

def GenerateConfigLoadHeader(bServer , filename , types , datas , comments , checks , css):
	filename = filename + g_loadConfigSuffix
	if GetCPPFilePath(bServer) == "":
		return
	outputPath = GetCPPFilePath(bServer) + os.sep + filename + ".h"
	if os.path.exists(outputPath): 
		os.remove(outputPath)

	fileWrite = open(outputPath , "a" , encoding='utf_8_sig')
	
	#以下是生成数据定义
	WriteFileDescription(fileWrite , filename + ".h" , "csv配置文件")
	fileWrite.write("#ifndef __" + filename + "_define_h__\n")
	fileWrite.write("#define __" + filename + "_define_h__\n") 
	fileWrite.write("#include \"CUtil/inc/Common.h \"\n") 
	fileWrite.write("#include \"Timer/inc/Date.h \"\n\n") 
	fileWrite.write("namespace " + g_xlsNamespace + "\n") 
	fileWrite.write("{\n") 
	fileWrite.write(oneTab + "struct " + g_configPrefix + filename + "\n") 
	fileWrite.write(oneTab + "{\n") 
	GenerateStructData(bServer , fileWrite , types , datas , comments , css , True)
	fileWrite.write(oneTab + "};\n\n\n") 
			
	#以下是生成导入数据接口
	fileWrite.write(oneTab + "class " + filename + "\n") 
	fileWrite.write(oneTab + "{\n") 
	fileWrite.write(oneTab + "public:\n")
	fileWrite.write(twoTab + "typedef std::vector<" + g_configPrefix + filename + "> CollectionConfigsT;\n\n")
	fileWrite.write(oneTab + "public:\n")
	fileWrite.write(twoTab + filename + "()\n")
	fileWrite.write(threeTab + ": m_bLoaded(false)\n")
	fileWrite.write(twoTab + "{}\n")
	fileWrite.write(oneTab + "\n")
	fileWrite.write(oneTab + "public:\n")
	fileWrite.write(twoTab + "bool				LoadFrom(const std::string& filename);\n\n")
	fileWrite.write(oneTab + "public:\n")
	fileWrite.write(twoTab + g_configPrefix + filename + " &	Get(size_t row);\n\n")
	fileWrite.write(oneTab + "public:\n")
	fileWrite.write(twoTab + "inline size_t		Count()")
	fileWrite.write("{ ")
	fileWrite.write("return m_vtConfigs.size(); ")
	fileWrite.write("}\n\n")

	#以下是生成导入数据动态检查和更改数据接口	
	fileWrite.write(oneTab + "public:\n")
	for index , item in enumerate(types):
		if not CheckCSType(css[index] , bServer):
			continue
		item_type = GetType(item)
		if item_type == g_structType or item_type == g_structArrayType:
			npos = datas[index].find('[')
			structName = datas[index][0 : npos]
			fileWrite.write(twoTab + "BOOL				" + g_checkFuncPrefix + structName + "(" + g_configPrefix + filename + " & conf);\n")			
		else:
			fileWrite.write(twoTab + "BOOL				" + g_checkFuncPrefix + datas[index] + "(" + g_configPrefix + filename + " & conf);\n")
	fileWrite.write(oneTab + "\n")

	fileWrite.write(oneTab + "private:\n")
	fileWrite.write(twoTab + "CollectionConfigsT	m_vtConfigs;\n")
	fileWrite.write(twoTab + "bool				m_bLoaded;\n")	
	fileWrite.write(oneTab + "};\n") 

	fileWrite.write("}\n\n" + "#endif// end" + "  __" + filename + "_define_h__\n") 

	fileWrite.close()	  
	
def GenerateConfigLoadCpp(bServer , filename , types , datas , comments , checks , css):
	loadFileName = filename + g_loadConfigSuffix
	if GetCPPFilePath(bServer) == "":
		return
	outputPath = GetCPPFilePath(bServer) + os.sep + loadFileName + ".cpp"
	if os.path.exists(outputPath): 
		os.remove(outputPath)

	fileWrite = open(outputPath , "a" , encoding='utf_8_sig')
	
	WriteFileDescription(fileWrite , loadFileName + ".cpp" , "csv读取文件实现")
	fileWrite.write("#include \"" + loadFileName + ".h\"\n") 
	fileWrite.write("#include \"../Condition.h\"\n") 
	fileWrite.write("#include \"CUtil/inc/CUtil.h\"\n") 
	fileWrite.write("#include \"CUtil/inc/CSVReader.h\"\n\n") 
	fileWrite.write("namespace " + g_xlsNamespace + "\n") 
	fileWrite.write("{\n") 
	fileWrite.write(oneTab + "bool " + loadFileName + "::LoadFrom(const std::string & filepath)\n") 
	fileWrite.write(oneTab + "{\n") 
	
	fileWrite.write(twoTab + "if (m_bLoaded)\n") 
	fileWrite.write(twoTab + "{\n") 
	fileWrite.write(threeTab + "return true;\n") 
	fileWrite.write(twoTab + "}\n") 
	
	fileWrite.write(twoTab + "CUtil::CSVReader csv;\n") 
	fileWrite.write(twoTab + "if(csv.Load(filepath) != 0)\n") 
	fileWrite.write(threeTab + "return false;\n\n") 

	for index , item in enumerate(types):
		if not CheckCSType(css[index] , bServer):
			continue
		item_type = GetType(item)
		if item_type == g_structType:
			npos = datas[index].find('[')
			structName = datas[index][0 : npos]
			fileWrite.write(twoTab + "size_t index_" + structName + " = csv.GetIndex(\"" + datas[index] + "\", " + str(g_rowName) + ");\n") 
			fileWrite.write(twoTab + "MsgAssert_Re0(index_" + structName + " != (size_t)-1 , \"error " + datas[index] + "\");\n\n") 
		elif item_type == g_structArrayType:
			npos = datas[index].find('[')
			structName = datas[index][0 : npos]
			fileWrite.write(twoTab + "size_t index_" + structName + " = csv.GetIndex(\"" + datas[index] + "\", " + str(g_rowName) + ");\n") 
			fileWrite.write(twoTab + "MsgAssert_Re0(index_" + structName + " != (size_t)-1 , \"error " + datas[index] + "\");\n\n") 
		else:
			fileWrite.write(twoTab + "size_t index_" + datas[index] + " = csv.GetIndex(\"" + datas[index] + "\", " + str(g_rowName) + ");\n") 
			fileWrite.write(twoTab + "MsgAssert_Re0(index_" + datas[index] + " != (size_t)-1 , \"error " + datas[index] + "\");\n\n") 

	fileWrite.write(twoTab + "for (size_t row = " + str(g_rowDataStart) + "; row < csv.Count(); ++row)\n") 
	fileWrite.write(twoTab + "{\n") 
	fileWrite.write(threeTab + g_configPrefix + loadFileName + " conf;\n\n") 
	for index , item in enumerate(types):
		if not CheckCSType(css[index] , bServer):
			continue
		GenerateLoadCppData(fileWrite , item , datas , index , loadFileName)

	#生成检查函数
	for index , item in enumerate(types):
		if not CheckCSType(css[index] , bServer):
			continue
		item_type = GetType(item)
		if item_type == g_structType or item_type == g_structArrayType:
			npos = datas[index].find('[')
			structName = datas[index][0 : npos]
			fileWrite.write(threeTab + "MsgAssert_Re0(" + g_checkFuncPrefix  + structName + "(conf) , \"" + structName + " check error.\");\n")			
		else:
			fileWrite.write(threeTab + "MsgAssert_Re0(" + g_checkFuncPrefix + datas[index] + "(conf) , \"" + datas[index] + " check error.\");\n")
	fileWrite.write(oneTab + "\n")

	fileWrite.write(threeTab + "m_vtConfigs.push_back(conf);\n") 
	fileWrite.write(twoTab + "}\n\n") 
	fileWrite.write(twoTab + "return true;\n") 
	fileWrite.write(oneTab + "}\n\n") 
			
	fileWrite.write(oneTab + g_configPrefix + loadFileName + " & " + loadFileName + "::Get(size_t row)\n") 
	fileWrite.write(oneTab + "{\n") 
	fileWrite.write(twoTab + "return m_vtConfigs.at(row);\n") 
	fileWrite.write(oneTab + "}\n\n") 

	GenerateLoadCppChecks(fileWrite , bServer , filename , types , datas , comments , checks , css)

	fileWrite.write(oneTab + "\n") 
	fileWrite.write("}\n\n") 
	
	fileWrite.close()	

def GenerateLoadCppChecks(fileWrite , bServer , filename , types , datas , comments , checks , css):
	for index , item in enumerate(types):
		if not CheckCSType(css[index] , bServer):
			continue
		GenerateLoadCppCheckFunc(fileWrite , bServer , filename , index , types , datas , checks)

def GenerateLoadCppCheckFunc(fileWrite , bServer , filename , index , types , datas , checks):
	loadFileName = filename + g_loadConfigSuffix
	item = types[index]
	item_type = GetType(item)
	if item_type == g_structType or item_type == g_structArrayType:
		npos = datas[index].find('[')
		structName = datas[index][0 : npos]
		fileWrite.write(oneTab + "BOOL	" + loadFileName + "::" + g_checkFuncPrefix + structName + "(" + g_configPrefix + loadFileName + " & conf)\n")			
	else:
		fileWrite.write(oneTab + "BOOL	" + loadFileName + "::" + g_checkFuncPrefix + datas[index] + "(" + g_configPrefix + loadFileName + " & conf)\n")

	fileWrite.write(oneTab + "{\n") 
	GenerateLoadCppCheckConditionFunc(fileWrite , bServer , filename , index , types , datas , checks)
	fileWrite.write(twoTab + "return TRUE;\n") 
	fileWrite.write(oneTab + "}\n") 

	fileWrite.write(oneTab + "\n")

def GenerateLoadCppCheckConditionFunc(fileWrite , bServer , filename , index , types , datas , checks):
	loadFileName = filename + g_loadConfigSuffix
	cellName = ""
	item = types[index]
	item_type = GetType(item)
	if item_type == g_structType or item_type == g_structArrayType:
		npos = datas[index].find('[')
		cellName = datas[index][0 : npos]
	else:
		cellName = datas[index]

	eachExpressions = None
	if bServer and filename in g_serverCellExpression and cellName in g_serverCellExpression[filename]:
		eachExpressions = g_serverCellExpression[filename][cellName]
	elif not bServer and filename in g_clientCellExpression and cellName in g_clientCellExpression[filename]:
		eachExpressions = g_clientCellExpression[filename][cellName]
	if eachExpressions != None and len(eachExpressions) > 0:		
		#LogOutDebug(":filename=" , filename , ":cellName=" , cellName , ":eachExpressions=" , eachExpressions)	
		for indexExpression , expression in enumerate(eachExpressions):
			objs = []
			GetExpressionObjects(expression , objs)

#			fileWrite.write(twoTab + "if (")
#			for objIndex , obj in enumerate(objs):
#				fileWrite.write(g_pointPrefix + obj.name + " != NULL")
#				if len(objs) - 1 != objIndex:				
#					fileWrite.write(" && ")
#				else:			
#					fileWrite.write(")\n")
#			fileWrite.write(twoTab + "{\n")

			fileWrite.write(twoTab + "if (" + expression.condition.text + ")\n")
			fileWrite.write(twoTab + "{\n")
			for actionIndex , action in enumerate(expression.actions):
				fileWrite.write(threeTab + action.name + "(" )
				for argIndex , arg in enumerate(action.args):	
					argName = arg
					#LogOutDebug("filename:", filename.upper() , " arg:" , arg) 
					for index , rowName in enumerate(g_xlsRecords[filename][g_rowName]): # 这里如果是表中的名字,则需要加上conf.
						if rowName.upper() == arg.upper():
							#LogOutDebug("rowName:", rowName.upper() , " arg:" , arg) 
							argName = "conf." + rowName + "" 
							break
								
					fileWrite.write(argName)
					if len(action.args) - 1 != argIndex:
						fileWrite.write(" , ")
				fileWrite.write(");\n")
			fileWrite.write(twoTab + "}\n")
#			fileWrite.write(twoTab + "}\n\n")

		#fileWrite.write(oneTab + "}\n\n") 


def GenerateLoadCppData(fileWrite , item , datas , index , filename):
	Str = GetTypeFunc(item)
	if type(Str) != str :
		if Str == g_int32ArrayFunc:
			fileWrite.write(threeTab + "{\n") 
			fileWrite.write(fourTab + "std::vector<std::string> vals;\n") 
			fileWrite.write(fourTab + "std::string __tmp = csv.GetString(row, index_" + datas[index] + ");\n") 
			fileWrite.write(fourTab + "CUtil::tokenize(__tmp, vals, \",\", \"\", \"\\\"\");\n") 
			fileWrite.write(fourTab + "for (size_t i = 0; i < vals.size(); ++i)\n") 
			fileWrite.write(fiveTab + "conf." + datas[index] + ".push_back((INT32)CUtil::atoi(vals[i].c_str()));\n") 
			fileWrite.write(threeTab + "}\n\n") 
		elif Str == g_int64ArrayFunc:
			fileWrite.write(threeTab + "{\n") 
			fileWrite.write(fourTab + "std::vector<std::string> vals;\n") 
			fileWrite.write(fourTab + "std::string __tmp = csv.GetString(row, index_" + datas[index] + ");\n") 
			fileWrite.write(fourTab + "CUtil::tokenize(__tmp, vals, \",\", \"\", \"\\\"\");\n") 
			fileWrite.write(fourTab + "for (size_t i = 0; i < vals.size(); ++i)\n") 
			fileWrite.write(fiveTab + "conf." + datas[index] + ".push_back((INT64)CUtil::atoi(vals[i].c_str()));\n") 
			fileWrite.write(threeTab + "}\n\n") 
		elif Str == g_boolArrayFunc:
			fileWrite.write(threeTab + "{\n") 
			fileWrite.write(fourTab + "std::vector<std::string> vals;\n") 
			fileWrite.write(fourTab + "std::string __tmp = csv.GetString(row, index_" + datas[index] + ");\n") 
			fileWrite.write(fourTab + "CUtil::tokenize(__tmp, vals, \",\", \"\", \"\\\"\");\n") 
			fileWrite.write(fourTab + "for (size_t i = 0; i < vals.size(); ++i)\n") 
			fileWrite.write(fiveTab + "conf." + datas[index] + ".push_back((bool)CUtil::atoi(vals[i].c_str()));\n") 
			fileWrite.write(threeTab + "}\n\n") 
		elif Str == g_dateArrayFunc:
			fileWrite.write(threeTab + "{\n") 
			fileWrite.write(fourTab + "std::vector<std::string> vals;\n") 
			fileWrite.write(fourTab + "std::string __tmp = csv.GetString(row, index_" + datas[index] + ");\n") 
			fileWrite.write(fourTab + "CUtil::tokenize(__tmp, vals, \",\", \"\", \"\\\"\");\n") 
			fileWrite.write(fourTab + "for (size_t i = 0; i < vals.size(); ++i)\n") 
			fileWrite.write(fiveTab + "conf." + datas[index] + ".push_back(Timer::Date(vals[i]));\n") 
			#fileWrite.write(fiveTab + "conf." + datas[index] + ".push_back(Timer::Date(vals[i], " + GetDateType(datas[index]) + "));\n") 
			fileWrite.write(threeTab + "}\n\n") 
		elif Str == g_doubleArrayFunc:
			fileWrite.write(threeTab + "{\n") 
			fileWrite.write(fourTab + "std::vector<std::string> vals;\n") 
			fileWrite.write(fourTab + "std::string __tmp = csv.GetString(row, index_" + datas[index] + ");\n") 
			fileWrite.write(fourTab + "CUtil::tokenize(__tmp, vals, \",\", \"\", \"\\\"\");\n") 
			fileWrite.write(fourTab + "for (size_t i = 0; i < vals.size(); ++i)\n") 
			fileWrite.write(fiveTab + "conf." + datas[index] + ".push_back((float)CUtil::atof(vals[i].c_str()));\n") 
			fileWrite.write(threeTab + "}\n\n")
		elif Str == g_stringArrayFunc:
			fileWrite.write(threeTab + "{\n") 
			fileWrite.write(fourTab + "std::string __tmp = csv.GetString(row, index_" + datas[index] + ");\n") 
			fileWrite.write(fourTab + "CUtil::tokenize(__tmp, conf." + datas[index] + ", \",\", \"\", \"\\\"\");\n") 
			fileWrite.write(threeTab + "}\n\n") 
		elif Str == g_structFunc:
			fileWrite.write(threeTab + "{\n") 

			npos = datas[index].find('[')
			structName = datas[index][0 : npos]
			structDatas = datas[index][npos + 1 : len(datas[index]) - 1].split(',')
			childItems = item.split(',')

			fileWrite.write(fourTab + "std::vector<std::string> vals;\n") 
			fileWrite.write(fourTab + "std::string __tmp = csv.GetString(row, index_" + structName + ");\n") 
			fileWrite.write(fourTab + "CUtil::tokenize(__tmp, vals, \",\", \"\", \"\\\"\");\n") 
			fileWrite.write(fourTab + "for (size_t i = 0; i < vals.size(); ++i)\n")
			fileWrite.write(fourTab + "{\n") 
			for indexChild , childItem in enumerate(childItems):
				fileWrite.write(fiveTab + "if(i == " + str(indexChild) + ")\n")
				fileWrite.write(fiveTab + "{\n") 

				strVal = GenerateStructDateByType(childItem , False)					
				
				fileWrite.write(sixTab + strVal + "\n") 
				fileWrite.write(sixTab + "conf." + structName + "." + structDatas[indexChild] + " = val;\n") 

				fileWrite.write(fiveTab + "}\n") 
			fileWrite.write(fourTab + "}\n") 
			fileWrite.write(threeTab + "}\n\n") 

		elif Str == g_structArrayFunc:		
			fileWrite.write(threeTab + "{\n") 

			npos = datas[index].find('[')
			structName = datas[index][0 : npos]
			structDatas = datas[index][npos + 1 : len(datas[index]) - 1].split(',')
			childItems = item.replace('[' , '').replace(']' , '').split(',')
			
			fileWrite.write(fourTab + "std::vector<std::string> vals;\n") 
			fileWrite.write(fourTab + "std::string __tmp = csv.GetString(row, index_" + structName + ");\n") 
			fileWrite.write(fourTab + "CUtil::tokenize(__tmp, vals, \"]\", \"\", \"\\\"\");\n") 
			fileWrite.write(fourTab + "for (size_t i = 0; i < vals.size(); ++i)\n")
			fileWrite.write(fourTab + "{\n") 
			fileWrite.write(fiveTab + "std::string strVal = vals[i];\n") 
			fileWrite.write(fiveTab + "if (strVal[0] == '[')\n") 
			fileWrite.write(sixTab + "strVal.assign(vals[i], 1, vals[i].length() - 1);\n\n") 
			fileWrite.write(fiveTab + g_configPrefix + filename + "::" + g_configPrefix + structName + "	" + "array;\n") 
			fileWrite.write(fiveTab + "std::vector<std::string> vals2;\n") 
			fileWrite.write(fiveTab + "CUtil::tokenize(strVal, vals2, \",\", \"\", \"\\\"\");\n") 
			fileWrite.write(fiveTab + "for (size_t j = 0; j < vals2.size(); ++j)\n") 
			fileWrite.write(fiveTab + "{\n")
			#childItems = re.split('(\[.*\])' , item)
			for indexChild , childItem in enumerate(childItems):
				fileWrite.write(sixTab + "if(j == " + str(indexChild) + ")\n")
				fileWrite.write(sixTab + "{\n") 

				strVal = GenerateStructDateByType(childItem , True)	
				
				fileWrite.write(sixTab + oneTab + strVal + "\n") 
				fileWrite.write(sixTab + oneTab  + "array." + structDatas[indexChild] + " = val;\n") 

				fileWrite.write(sixTab + "}\n") 
			fileWrite.write(fiveTab + "}\n") 
			fileWrite.write(fiveTab + "conf.vec" + structName + ".push_back(array);\n") 
			fileWrite.write(fourTab + "}\n") 
			fileWrite.write(threeTab + "}\n\n")				 
		elif Str == g_mapFunc:
			fileWrite.write(threeTab + "{\n") 
			fileWrite.write(fourTab + "std::vector<std::string> vals;\n") 
			fileWrite.write(fourTab + "std::string __tmp = csv.GetString(row, index_" + datas[index] + ");\n") 
			fileWrite.write(fourTab + "CUtil::tokenize(__tmp, vals, \"=\", \"\", \"\\\"\");\n") 
			fileWrite.write(fourTab + "for (size_t i = 0; i < vals.size(); ++i)\n") 
			
			mapType , keyType , valueType = GetMapType(item)
			keyInput = GenerateMapDateByType(keyType , True)
			valueInput = GenerateMapDateByType(valueType , False)
			
			fileWrite.write(fiveTab + "conf." + datas[index] + ".insert(std::make_pair(" + keyInput + "," + valueInput + "));\n") 
			fileWrite.write(threeTab + "}\n\n") 
		elif Str == g_configFunc:
			tmp = collections.OrderedDict()
			MakeDicKeyUpper(g_xlsRecords , tmp)
			parentName = GetDicKeyByUpper(g_xlsRecords , item)
			itemType , configKeyType , configValueType = GetConfigType(item)
			
			GenerateLoadCppData(fileWrite , configKeyType , datas , index , filename)
	else:
		fileWrite.write(threeTab + "conf." + datas[index] + " = csv." + Str + "(row , index_" + datas[index] + ");\n") 

def GenerateMapDateByType(itemType , isKey):
	keyWord = ""
	result = ""
	if isKey:
		result = "vals[0]"
		keyWord = "vals[0]"
	else:
		result = "vals[1]"
		keyWord = "vals[1]"
		
	if itemType == g_boolType:
		result = "!!" + "CUtil::atoi(" + keyWord + ")"
	if itemType == g_int32Type:
		result = "(" + g_int32Type + ")CUtil::atoi(" + keyWord + ")"
	if itemType == g_int64Type:
		result = "(" + g_int64Type + ")CUtil::atoi(" + keyWord + ")"
	if itemType == g_doubleType:
		result = "(" + g_doubleType + ")CUtil::atof(" + keyWord + ")"
	if itemType == g_dateType:
		result = "Timer::Date(" + keyWord + ")"
	if itemType == g_configType:
		tmp = collections.OrderedDict()
		MakeDicKeyUpper(g_xlsRecords , tmp)
		parentName = GetDicKeyByUpper(g_xlsRecords , item)
		itemType , configKeyType , configValueType = GetConfigType(item)
		
		return GenerateMapDateByType(GetType(configKeyType) , isKey)
		
	return result
	
def GenerateStructDateByType(childItem , bArray):
	strVal = ""
	strVals = ""
	if bArray == True:
		strVals = "vals2[j]"
	else:
		strVals = "vals[i]"		
	
	childItemType = GetType(childItem)
	if childItemType == g_boolType:
		strVal = "bool val = "
		strVal = strVal + "CUtil::strtobool(" + strVals + ".c_str()) >= 1;"
	elif childItemType == g_int32Type:
		strVal = "INT32 val = "
		strVal = strVal + "(INT32)CUtil::atoi(" + strVals + ".c_str());"
	elif childItemType == g_dateType:
		strVal = "Timer::Date val(vals[i]);"
	elif childItemType == g_int64Type:
		strVal = "INT64 val = "
		strVal = strVal + "(INT64)CUtil::atoi(" + strVals + ".c_str());"
	elif childItemType == g_doubleType:
		strVal = "double val = "
		strVal = strVal + "(float)CUtil::atof(" + strVals + ".c_str());"
	elif childItemType == g_stringType:
		strVal = "std::string val = "
		strVal = strVal + strVals + ".c_str();"
	elif childItemType == g_configType:
		tmp = collections.OrderedDict()
		MakeDicKeyUpper(g_xlsRecords , tmp)
		parentName = GetDicKeyByUpper(g_xlsRecords , childItem)
		itemType , keyType , valueType = GetConfigType(childItem)
		
		return GenerateStructDateByType(keyType , bArray)
		
	return strVal

def GenerateConfigDeclareHeader(bServer): 
	if GetCPPFilePath(bServer) == "":
		return
	outputPath = GetCPPFilePath(bServer) + os.sep + "ConfigDeclare.h"
	if os.path.exists(outputPath): 
		os.remove(outputPath) 

	fileWrite = open(outputPath , "a" , encoding='utf_8_sig') 

	WriteFileDescription(fileWrite , "ConfigDeclare.h" , "ConfigDeclare")
	fileWrite.write("#ifndef __" + g_xlsNamespace + "_ConfigDeclare" + "_define_h__\n")
	fileWrite.write("#define __" + g_xlsNamespace + "_ConfigDeclare" +  "_define_h__\n")  
	fileWrite.write("namespace " + g_xlsNamespace + "\n") 
	fileWrite.write("{\n\n") 
			
	for sheet , item in g_xlsRecords.items():
		fileWrite.write(oneTab + "struct " + g_configPrefix + sheet + " ;\n")  

	fileWrite.write("}\n\n" + "#endif// end" + "  __" + g_xlsNamespace + "_ConfigDeclare_define_h__\n") 
	
	fileWrite.close()	 	
	
def GenerateConfigBaseHeader(bServer , filename , types , datas , comments , css):
	if GetCPPFilePath(bServer) == "":
		return
	newFileName = filename + g_configSuffix
	outputPath = GetCPPFilePath(bServer) + os.sep + newFileName + ".h"

	if os.path.exists(outputPath): 
		os.remove(outputPath)

	fileWrite = open(outputPath , "a" , encoding='utf_8_sig')
	
	loadConfig = filename + g_loadConfigSuffix
	dataConfig = g_configPrefix + filename

	strTitle = MakeTitle(types , datas)
	fileWrite.write(strTitle)	
	WriteFileDescription(fileWrite , newFileName + ".h" , "csv读取文件")
	fileWrite.write("#ifndef __" + g_xlsNamespace + "_" + newFileName + "_define_h__\n")
	fileWrite.write("#define __" + g_xlsNamespace + "_" + newFileName +  "_define_h__\n") 
	fileWrite.write("#include \"" + loadConfig + ".h\"\n") 
	fileWrite.write("#include \"../Condition.h\"\n") 
	fileWrite.write("#include \"ConfigDeclare.h\"\n") 
	fileWrite.write("#include \"CUtil/inc/CSVConfig.h\"\n\n") 

	fileWrite.write("namespace " + g_xlsNamespace + "\n") 
	fileWrite.write("{\n\n") 

	fileWrite.write(oneTab + "struct " + dataConfig + "\n") 
	fileWrite.write(oneTab + "{\n") 	 
	GenerateStructData(bServer , fileWrite , types , datas , comments , css , False)
	fileWrite.write(oneTab + "};\n\n\n") 
			
	fileWrite.write(oneTab + "class " + newFileName + ": public CUtil::CSVConfig\n") 
	fileWrite.write(oneTab + "{\n") 
	
	fileWrite.write(oneTab + "public:\n")
	fileWrite.write(twoTab + "typedef std_unordered_map<" + GetType(types[0]) + " , " + dataConfig + "> MapConfigsT;\n\n")
	fileWrite.write(oneTab + "public:\n")
	fileWrite.write(twoTab + "bool" + fourTab + "LoadFrom(const std::string& filepath);\n")
	fileWrite.write(twoTab + "bool" + fourTab + "RepairLoad(const std::string& filepath);\n")
	fileWrite.write(twoTab + dataConfig + oneTab + "*" + oneTab + "Get" + filename + "(" + GetType(g_xlsRecords[filename][g_rowType][g_cellID]) + " id , std::string strFilePath = \"\");\n\n")
	GenerateConditionFunc(fileWrite , bServer , filename , types , datas , comments , css)
	fileWrite.write(oneTab + "private:\n")
	fileWrite.write(twoTab + "MapConfigsT			m_mapConfigs;\n\n")
	
	fileWrite.write(oneTab + "};\n") 


	fileWrite.write("}\n\n" + "#endif// end" + "  __" + g_xlsNamespace + "_" + filename + "_define_h__\n") 
	
	fileWrite.close()	 	

def GenerateConditionFunc(fileWrite , bServer , filename , types , datas , comments , css):	
	eachExpressions = None
	if bServer and filename in g_serverExpression:
		eachExpressions = g_serverExpression[filename]
	elif not bServer and filename in g_clientExpression:
		eachExpressions = g_clientExpression[filename]
	if eachExpressions != None:			
		fileWrite.write(oneTab + "public:\n")
		for index , expressions in eachExpressions.items():
			GenerateExpressionFuncName(False , fileWrite , filename , index , expressions)
		fileWrite.write("\n\n")

def GenerateExpressionFuncName(bCPP , fileWrite , filename , index , expressions):
	newFileName = filename + g_configSuffix
	objs = []
	GetExpressionsObjects(expressions , objs)
	if not bCPP:
		fileWrite.write(twoTab + "bool" + fourTab + index + "(" + GetType(g_xlsRecords[filename][g_rowType][g_cellID]) + " id")
	else:
		fileWrite.write(oneTab + "bool " + newFileName + "::" + index + "(" + GetType(g_xlsRecords[filename][g_rowType][g_cellID]) + " id")
		
	for objIndex , obj in enumerate(objs):
		if len(obj.name) > 0:
			fileWrite.write(" , ")
			for objNamespace , namespace in enumerate(obj.namespace):
				fileWrite.write(namespace)
				fileWrite.write("::")
			fileWrite.write(obj.name) 
			if not bCPP:
				fileWrite.write(" * " + g_pointPrefix + obj.name + " = NULL") 
			else:
				fileWrite.write(" * " + g_pointPrefix + obj.name + "/* = NULL*/" ) 
		
	if not bCPP:
		fileWrite.write(");\n")
	else:
		fileWrite.write(")\n")
		
def GenerateConfigBaseCpp(bServer , filename , types , datas , comments , css):
	if GetCPPFilePath(bServer) == "":
		return

	newFileName = filename + g_configSuffix
	outputPath = GetCPPFilePath(bServer) + os.sep + newFileName + ".cpp"
	if os.path.exists(outputPath): 
		os.remove(outputPath)

	loadConfig = filename + g_loadConfigSuffix
	dataConfig = g_configPrefix + filename
	fileWrite = open(outputPath , "a" , encoding='utf_8_sig')
	
	strTitle = MakeTitle(types , datas)
	fileWrite.write(strTitle)	
	WriteFileDescription(fileWrite , newFileName + ".cpp" , "csv读取数据文件实现")
	fileWrite.write("#include \"" + newFileName + ".h\"\n") 
	fileWrite.write("#include \"LogLib/inc/Log.h\"\n") 

	GenerateConfigHeaderInlude(fileWrite , bServer , filename , types , datas , comments , css , newFileName , loadConfig , dataConfig)
							
	fileWrite.write("namespace " + g_xlsNamespace + "\n") 
	fileWrite.write("{\n") 
	
	GenerateConfigBaseCppLoadFrom(fileWrite , bServer , filename , types , datas , comments , css , newFileName , loadConfig, dataConfig)
	GenerateConfigBaseCppRepairLoad(fileWrite , bServer , filename , types , datas , comments , css , newFileName , loadConfig, dataConfig)
	
	fileWrite.write(oneTab + dataConfig + " * " + newFileName + "::Get" + filename + "(" + GetType(g_xlsRecords[filename][g_rowType][g_cellID]) + " id , std::string strFilePath/* = \"\"*/)\n")
	fileWrite.write(oneTab + "{\n") 
	#fileWrite.write(twoTab + "if (!m_bLoaded)\n") 
	#fileWrite.write(twoTab + "{\n") 
	#fileWrite.write(threeTab + "LoadFrom(strFilePath);\n") 
	#fileWrite.write(twoTab + "}\n") 
	fileWrite.write(twoTab + "MapConfigsT::iterator iter = m_mapConfigs.find(id);\n") 
	fileWrite.write(twoTab + "if(iter == m_mapConfigs.end())\n") 
	fileWrite.write(twoTab + "{\n") 
#	fileWrite.write(threeTab + "gWarniStream( \"" + filename + "::Get" + filename + " NotFound \" << id);\n") 
	fileWrite.write(threeTab + "return NULL;\n") 
	fileWrite.write(twoTab + "}\n\n") 
	fileWrite.write(twoTab + "return &iter->second;\n") 
	fileWrite.write(oneTab + "}\n\n") 

	GenerateConditionCppFunc(fileWrite , bServer , filename , types , datas , comments , css)

	fileWrite.write("}\n\n") 
	
	fileWrite.close()

def GenerateConfigHeaderInlude(fileWrite , bServer , filename , types , datas , comments , css , newFileName , loadConfig , dataConfig):
	sameParaentName = collections.OrderedDict()
	for index , item in enumerate(types):
		if not CheckCSType(css[index] , bServer):
			continue
		itemType = GetType(item)
		if itemType == g_structType:
			npos = datas[index].find('[')
			structName = datas[index][0 : npos]
			structDatas = datas[index][npos + 1 : len(datas[index]) - 1].split(',')
			childItems = item.split(',')

			for indexChild , childItem in enumerate(childItems):
				childItem = RemoveSpecialWord(childItem)
				if GetType(childItem) == g_configType:
					tmp = collections.OrderedDict()
					MakeDicKeyUpper(g_xlsRecords , tmp)
					parentName = GetDicKeyByUpper(g_xlsRecords , childItem)

					if parentName not in sameParaentName:
						sameParaentName[parentName] = parentName
						fileWrite.write("#include \"" + parentName + ".h\"\n")
						
		elif itemType == g_structArrayType:			
			npos = datas[index].find('[')
			structName = datas[index][0 : npos]
			structDatas = datas[index][npos + 1 : len(datas[index]) - 1].split(',')
			childItems = item.split(',')

			for indexChild , childItem in enumerate(childItems):
				childItem = RemoveSpecialWord(childItem)
				if GetType(childItem) == g_configType:
					tmp = collections.OrderedDict()
					MakeDicKeyUpper(g_xlsRecords , tmp)
					parentName = GetDicKeyByUpper(g_xlsRecords , childItem)

					if parentName not in sameParaentName:
						sameParaentName[parentName] = parentName
						fileWrite.write("#include \"" + parentName + ".h\"\n")
						
		elif itemType == g_configType:
			tmp = collections.OrderedDict()
			MakeDicKeyUpper(g_xlsRecords , tmp)
			parentName = GetDicKeyByUpper(g_xlsRecords , item)
					
			if parentName not in sameParaentName:
				sameParaentName[parentName] = parentName
				fileWrite.write("#include \"" + parentName + ".h\"\n")
	fileWrite.write("\n") 
	
def GenerateConfigBaseCppLoadFrom(fileWrite , bServer , filename , types , datas , comments , css , newFileName , loadConfig , dataConfig):
	fileWrite.write(oneTab + "bool " + newFileName + "::LoadFrom(const std::string & filepath)\n") 
	fileWrite.write(oneTab + "{\n") 
	
	fileWrite.write(twoTab + "if (m_bLoaded)\n") 
	fileWrite.write(twoTab + "{\n") 
	fileWrite.write(threeTab + "return false;\n") 
	fileWrite.write(twoTab + "}\n") 
	fileWrite.write(twoTab + g_xlsNamespace + "::" + loadConfig + " loadConfig;\n") 
	fileWrite.write(twoTab + "MsgAssert_Re0(loadConfig.LoadFrom(filepath + " + "\"" + filename + g_csvFileSuffix + "\"" + ") , \"Error " + filename + "LoadFrom \" << filepath + " + "\"" + filename + g_csvFileSuffix + "\"" + ");\n\n") 
	
	fileWrite.write(twoTab + "for(size_t i = 0; i < loadConfig.Count(); ++i)\n") 
	fileWrite.write(twoTab + "{\n") 
	fileWrite.write(threeTab + g_xlsNamespace + "::" + g_configPrefix + loadConfig + "& config = loadConfig.Get(i);\n") 
	fileWrite.write(threeTab + g_xlsNamespace + "::" + dataConfig + " data;\n") 
	
	for index , item in enumerate(types):
		if not CheckCSType(css[index] , bServer):
			continue
		itemType = GetType(item)
		if itemType == g_structType:
			npos = datas[index].find('[')
			structName = datas[index][0 : npos]
			structDatas = datas[index][npos + 1 : len(datas[index]) - 1].split(',')
			childItems = item.split(',')

			fileWrite.write(threeTab + "{\n") 
			for indexChild , childItem in enumerate(childItems):
				childItem = RemoveSpecialWord(childItem)
				if GetType(childItem) == g_configType:
					tmp = collections.OrderedDict()
					MakeDicKeyUpper(g_xlsRecords , tmp)
					parentName = GetDicKeyByUpper(g_xlsRecords , childItem)

					fileWrite.write(fourTab + "data." + structName + "." + structDatas[indexChild] + ".insert(std::make_pair(config." + structName + "." + structDatas[indexChild] + " , " + g_xlsNamespace + "::" + parentName + "::GetInstance().Get" + parentName + "(config." + structName + "." + structDatas[indexChild] + " , filepath)));\n") 
				else:
					fileWrite.write(fourTab + "data." + structName + "." + structDatas[indexChild] + " = config." + structName + "." + structDatas[indexChild] + ";\n") 
			fileWrite.write(threeTab + "}\n")

		elif itemType == g_structArrayType:			
			npos = datas[index].find('[')
			structName = datas[index][0 : npos]
			structDatas = datas[index][npos + 1 : len(datas[index]) - 1].split(',')
			childItems = item.split(',')

			fileWrite.write(threeTab + "{\n") 
			fileWrite.write(fourTab + "std::vector<" + g_configPrefix + loadConfig + "::" + g_configPrefix + structName + ">::iterator iter = config.vec" + structName + ".begin();\n") 
			fileWrite.write(fourTab + "std::vector<" + g_configPrefix + loadConfig + "::" + g_configPrefix + structName + ">::iterator end = config.vec" + structName + ".end();\n") 
			fileWrite.write(fourTab + "for (; iter != end;++iter)\n") 
			fileWrite.write(fourTab + "{\n") 
			fileWrite.write(fiveTab + g_configPrefix + filename + "::" + g_configPrefix + structName + " array;\n")
			for indexChild , childItem in enumerate(childItems):
				childItem = RemoveSpecialWord(childItem)
				if GetType(childItem) == g_configType:
					tmp = collections.OrderedDict()
					MakeDicKeyUpper(g_xlsRecords , tmp)
					parentName = GetDicKeyByUpper(g_xlsRecords , childItem)

					fileWrite.write(fiveTab + "array." + structDatas[indexChild] + ".insert(std::make_pair(iter->" + structDatas[indexChild] + " , "+ g_xlsNamespace + "::" + parentName + "::GetInstance().Get" + parentName + "(iter->" + structDatas[indexChild] + " , filepath)));\n") 
				else:
					fileWrite.write(fiveTab + "array." + structDatas[indexChild] + " = iter->" + structDatas[indexChild] + ";\n")
			fileWrite.write(fiveTab + "data.vec" + structName + ".push_back(array);\n")
			fileWrite.write(fourTab + "}\n")
			fileWrite.write(threeTab + "}\n")
		elif itemType == g_configType:
			tmp = collections.OrderedDict()
			MakeDicKeyUpper(g_xlsRecords , tmp)
			parentName = GetDicKeyByUpper(g_xlsRecords , item)

			fileWrite.write(threeTab + "data." + datas[index] + ".insert(std::make_pair(config." + datas[index] + " , " + g_xlsNamespace + "::" + parentName + "::GetInstance().Get" + parentName + "(config." + datas[index] + " , filepath)));\n") 
		else:
			fileWrite.write(threeTab + "data." + datas[index] + " = config." + datas[index] + ";\n") 

	fileWrite.write(threeTab + "m_mapConfigs.insert(std::make_pair(data." + datas[0] + ",data));\n") 
	fileWrite.write(twoTab + "}\n\n") 
	fileWrite.write(twoTab + "m_bLoaded = true;\n") 
	fileWrite.write(twoTab + "return true;\n") 
	fileWrite.write(oneTab + "}\n\n") 
			
	
def GenerateConfigBaseCppRepairLoad(fileWrite , bServer , filename , types , datas , comments , css, newFileName , loadConfig, dataConfig):
	fileWrite.write(oneTab + "bool " + newFileName + "::RepairLoad(const std::string & filepath)\n") 
	fileWrite.write(oneTab + "{\n") 
	
	fileWrite.write(twoTab + "MsgAssert_Re0(m_bLoaded , \"error " + filename + " .no load data.\")\n") 
	
	fileWrite.write(twoTab + "MapConfigsT::iterator iterConfig = m_mapConfigs.begin();\n\n") 	
	fileWrite.write(twoTab + "for(; iterConfig != m_mapConfigs.end(); ++iterConfig)\n") 
	fileWrite.write(twoTab + "{\n") 
	fileWrite.write(threeTab + g_xlsNamespace + "::" + dataConfig + " & data = iterConfig->second;\n") 
	
	for index , item in enumerate(types):
		if not CheckCSType(css[index] , bServer):
			continue
		itemType = GetType(item)
		if itemType == g_structType:
			npos = datas[index].find('[')
			structName = datas[index][0 : npos]
			structDatas = datas[index][npos + 1 : len(datas[index]) - 1].split(',')
			childItems = item.split(',')

			fileWrite.write(threeTab + "{\n") 
			for indexChild , childItem in enumerate(childItems):
				childItem = RemoveSpecialWord(childItem)
				if GetType(childItem) == g_configType:
					tmp = collections.OrderedDict()
					MakeDicKeyUpper(g_xlsRecords , tmp)
					parentName = GetDicKeyByUpper(g_xlsRecords , childItem)

					fileWrite.write(fourTab + "data." + structName + "." + structDatas[indexChild] + ".begin()->second = " + g_xlsNamespace + "::" + parentName + "::GetInstance().Get" + parentName + "(data." + structName + "." + structDatas[indexChild] + ".begin()->first" + " , filepath);\n") 
				#else:
				#	fileWrite.write(fourTab + "data." + structName + "." + structDatas[indexChild] + " = config." + structName + "." + structDatas[indexChild] + ";\n") 
			fileWrite.write(threeTab + "}\n")

		elif itemType == g_structArrayType:			
			npos = datas[index].find('[')
			structName = datas[index][0 : npos]
			structDatas = datas[index][npos + 1 : len(datas[index]) - 1].split(',')
			childItems = item.split(',')

			fileWrite.write(threeTab + "{\n") 
			fileWrite.write(fourTab + "std::vector<" + dataConfig + "::" + g_configPrefix + structName + "> & vec" + structName + " = data.vec" + structName + ";\n") 
			fileWrite.write(fourTab + "std::vector<" + dataConfig + "::" + g_configPrefix + structName + ">::iterator iter" + structName + " = vec" + structName + ".begin();\n") 
			fileWrite.write(fourTab + "std::vector<" + dataConfig + "::" + g_configPrefix + structName + ">::iterator end" + structName + " = vec" + structName + ".end();\n") 
			fileWrite.write(fourTab + "for (; iter" + structName + " != end" + structName + ";++iter" + structName + ")\n") 
			fileWrite.write(fourTab + "{\n") 
			fileWrite.write(fiveTab + g_configPrefix + filename + "::" + g_configPrefix + structName + " & array = *iter" + structName + ";\n")
			for indexChild , childItem in enumerate(childItems):
				childItem = RemoveSpecialWord(childItem)
				if GetType(childItem) == g_configType:
					tmp = collections.OrderedDict()
					MakeDicKeyUpper(g_xlsRecords , tmp)
					parentName = GetDicKeyByUpper(g_xlsRecords , childItem)

					fileWrite.write(fiveTab + "array." + structDatas[indexChild] + ".begin()->second = " + g_xlsNamespace + "::" + parentName + "::GetInstance().Get" + parentName + "(array." + structDatas[indexChild] + ".begin()->first" + " , filepath);\n") 
				#else:
				#	fileWrite.write(fiveTab + "array." + structDatas[indexChild] + " = iter->" + structDatas[indexChild] + ";\n")
			#fileWrite.write(fiveTab + "data.vec" + structName + ".push_back(array);\n")
			fileWrite.write(fourTab + "}\n")
			fileWrite.write(threeTab + "}\n")
		elif itemType == g_configType:
			tmp = collections.OrderedDict()
			MakeDicKeyUpper(g_xlsRecords , tmp)
			parentName = GetDicKeyByUpper(g_xlsRecords , item)
					
			fileWrite.write(threeTab + "data." + datas[index] + ".begin()->second = " + g_xlsNamespace + "::" + parentName + "::GetInstance().Get" + parentName + "(data." + datas[index] + ".begin()->first" + " , filepath);\n") 
		#else:
		#	fileWrite.write(threeTab + "data." + datas[index] + " = config." + datas[index] + ";\n") 

	fileWrite.write(twoTab + "}\n\n") 
	fileWrite.write(twoTab + "return true;\n") 
	fileWrite.write(oneTab + "}\n\n") 
			
def GenerateConditionCppFunc(fileWrite , bServer , filename , types , datas , comments , css):	
	eachExpressions = None
	if bServer and filename in g_serverExpression:
		eachExpressions = g_serverExpression[filename]
	elif (not bServer) and filename in g_clientExpression:
		eachExpressions = g_clientExpression[filename]
	if eachExpressions != None and len(eachExpressions) > 0:			
		for index , expressions in eachExpressions.items():
			GenerateExpressionFuncName(True , fileWrite , filename , index , expressions)

			fileWrite.write(oneTab + "{\n")
			fileWrite.write(twoTab + "if (Get" + filename +"(id) == NULL)\n")
			fileWrite.write(twoTab + "{\n")
			fileWrite.write(threeTab + "gErrorStream(\"" + g_expressionPrefix + "Use error. " + filename + "  no this id.id=\" << id);\n")
			fileWrite.write(threeTab + "return false;\n")
			fileWrite.write(twoTab + "}\n\n")

			for indexExpression , expression in enumerate(expressions):
				objs = []
				GetExpressionObjects(expression , objs)

				fileWrite.write(twoTab + "if (")
				for objIndex , obj in enumerate(objs):
					fileWrite.write(g_pointPrefix + obj.name + " != NULL")
					if len(objs) - 1 != objIndex:				
						fileWrite.write(" && ")
					else:			
						fileWrite.write(")\n")
				fileWrite.write(twoTab + "{\n")

				fileWrite.write(threeTab + "if (" + expression.condition.text + ")\n")
				fileWrite.write(threeTab + "{\n")
				for actionIndex , action in enumerate(expression.actions):
					fileWrite.write(fourTab + action.name + "(" )
					for argIndex , arg in enumerate(action.args):
						fileWrite.write(arg)
						if len(action.args) - 1 != argIndex:
							fileWrite.write(" , ")
					fileWrite.write(");\n")
				fileWrite.write(threeTab + "}\n")
				fileWrite.write(twoTab + "}\n\n")

			fileWrite.write(twoTab + "return true;\n")
			fileWrite.write(oneTab + "}\n\n") 

def GenerateConfigHeader(bServer , filename , types , datas , comments , css):
	if GetCPPFilePath(bServer) == "":
		return
	outputPath = GetCPPFilePath(bServer) + os.sep + filename + ".h"
	#if CheckNeedDelete(outputPath , types , datas ):
	#	os.remove(outputPath)
		#return
	if os.path.exists(outputPath):
		os.remove(outputPath)
		#return

	fileWrite = open(outputPath , "a" , encoding='utf_8_sig')
	
	loadConfig = filename + g_loadConfigSuffix
	dataConfig = g_configPrefix + filename

	strTitle = MakeTitle(types , datas)
	fileWrite.write(strTitle)	
	WriteFileDescription(fileWrite , filename + ".h" , "csv读取文件")
	fileWrite.write("#ifndef __" + g_xlsNamespace + "_" + filename + "_define_h__\n")
	fileWrite.write("#define __" + g_xlsNamespace + "_" + filename +  "_define_h__\n") 
	#fileWrite.write("#include \"" + loadConfig + ".h\"\n") 
	#fileWrite.write("#include \"../Condition.h\"\n") 
	#fileWrite.write("#include \"CUtil/inc/CSVConfig.h\"\n\n") 
	fileWrite.write("#include \"" + filename + g_configSuffix + ".h\"\n\n") 


	fileWrite.write("namespace " + g_xlsNamespace + "\n") 
	fileWrite.write("{\n\n") 

	fileWrite.write(oneTab + "class " + filename + ": public " + filename + g_configSuffix + "\n") 
	fileWrite.write(oneTab + "{\n") 
	
	fileWrite.write(oneTab + "public:\n")
	fileWrite.write(twoTab + "static " + filename + "&	GetInstance(){ static " + filename + " s_" + filename + "; return s_" + filename + "; }\n")
	fileWrite.write(twoTab + "static " + filename + "*	GetInstancePtr(){ return &GetInstance(); }\n")
	fileWrite.write(twoTab + "\n")
	
	fileWrite.write(oneTab + "public:\n")
	fileWrite.write(twoTab + "virtual BOOL		OnLoad();\n")
	fileWrite.write(twoTab + "\n")
	fileWrite.write(oneTab + "private:\n\n")
	
	fileWrite.write(oneTab + "};\n") 			
#	fileWrite.write(oneTab + "extern " + filename + " * " + g_globaleNamePrefix + filename + ";\n") 


	fileWrite.write("}\n\n" + "#endif// end" + "  __" + g_xlsNamespace + "_" + filename + "_define_h__\n") 
	
	fileWrite.close()	 	

def GenerateConfigCpp(bServer , filename , types , datas , comments , css):
	if GetCPPFilePath(bServer) == "":
		return
	outputPath = GetCPPFilePath(bServer) + os.sep + filename + ".cpp"
	#if CheckNeedDelete(outputPath , types , datas ):
	#	os.remove(outputPath)
		#return
	if os.path.exists(outputPath):
		os.remove(outputPath)
		#return

	loadConfig = filename + g_loadConfigSuffix
	dataConfig = g_configPrefix + filename
	fileWrite = open(outputPath , "a" , encoding='utf_8_sig')
	
	strTitle = MakeTitle(types , datas)
	fileWrite.write(strTitle)	
	WriteFileDescription(fileWrite , filename + ".cpp" , "csv读取数据文件实现")
	fileWrite.write("#include \"" + filename + ".h\"\n") 
	fileWrite.write("#include \"LogLib/inc/Log.h\"\n") 

	GenerateConfigHeaderInlude(fileWrite , bServer , filename , types , datas , comments , css , "" , loadConfig , dataConfig)
	
	fileWrite.write("namespace " + g_xlsNamespace + "\n") 
	fileWrite.write("{\n") 
#	fileWrite.write(oneTab + filename + " * " + g_globaleNamePrefix + filename + " = NULL;\n\n") 
	fileWrite.write(oneTab + "//tools after data load success , call OnLoad;\n") 
	fileWrite.write(oneTab + "BOOL " + filename + "::OnLoad()\n") 
	fileWrite.write(oneTab + "{\n") 
	fileWrite.write(twoTab + "return FALSE;\n") 
	fileWrite.write(oneTab + "}\n") 
	fileWrite.write("}\n\n") 
	
	fileWrite.close()

def GetExpressionsObjects(expressions , objs):
	for indexExpression , expression in enumerate(expressions):
		GetExpressionObjects(expression , objs)

def GetExpressionObjects(expression , objs):
	#LogOutDebug("expression:" , expression)
	for conditionIndex , conditionObj in enumerate(expression.condition.objs):
		bIn = False
		for index , obj in enumerate(objs):
			if obj.name == conditionObj.name:
				bIn = True
		if not bIn:		
			#LogOutDebug("conditionObj:" , conditionObj.name)
			objs.append(conditionObj)
	for actionIndex , action in enumerate(expression.actions):
		for objIndex , actionObj in enumerate(action.objs):
			bIn = False
			for index , obj in enumerate(objs):
				if obj.name == actionObj.name:
					bIn = True
			if not bIn:
				objs.append(actionObj)

def GenerateConfigManagerHeader(bServer):
	if GetCPPFilePath(bServer) == "":
		return
	outputPath = GetCPPFilePath(bServer) + os.sep + "ConfigManager.h"
	if os.path.exists(outputPath): 
		os.remove(outputPath) 

	fileWrite = open(outputPath , "a" , encoding='utf_8_sig') 

	WriteFileDescription(fileWrite , "ConfigManager.h" , "ConfigManager文件声明")
	fileWrite.write("#ifndef __" + g_xlsNamespace + "_ConfigManager" + "_define_h__\n")
	fileWrite.write("#define __" + g_xlsNamespace + "_ConfigManager" +  "_define_h__\n")
	fileWrite.write("#include \"CUtil/inc/Common.h\"\n")  
	fileWrite.write("#include \"Lua/lua_tinker.h\"\n\n") 
	fileWrite.write("namespace " + g_xlsNamespace + "\n") 
	fileWrite.write("{\n\n") 
			
	fileWrite.write(oneTab + "class ConfigManager\n") 
	fileWrite.write(oneTab + "{\n") 
	
	fileWrite.write(oneTab + "public:\n")
	fileWrite.write(twoTab + "ConfigManager();\n")
	fileWrite.write(twoTab + "virtual ~ConfigManager();\n\n")
	fileWrite.write(oneTab + "public:\n")
	fileWrite.write(twoTab + "static ConfigManager & GetInstance();\n\n")
	fileWrite.write(oneTab + "public:\n")
	fileWrite.write(twoTab + "virtual INT32		Init(std::string  strCsvPath);\n")
	fileWrite.write(twoTab + "virtual INT32		Cleanup();\n")
	fileWrite.write(twoTab + "virtual INT32		LoadFrom(std::string  strCsvPath);\n")
	fileWrite.write(twoTab + "virtual INT32		RepairLoad(std::string  strCsvPath);\n")
	fileWrite.write(twoTab + "virtual INT32		ExportClassToLua(lua_State* L);\n")
	fileWrite.write(twoTab + "\n")
	fileWrite.write(oneTab + "public:\n")
	fileWrite.write(oneTab + "private:\n")
	fileWrite.write(twoTab + "\n\n")
	fileWrite.write(oneTab + "};\n") 		


	fileWrite.write("}\n\n" + "#endif// end" + "  __" + g_xlsNamespace + "_ConfigManager_define_h__\n") 
	
	fileWrite.close()	 	


def GenerateConfigManagerCPP(bServer):
	if GetCPPFilePath(bServer) == "":
		return
	outputPath = GetCPPFilePath(bServer) + os.sep + "ConfigManager.cpp"
	if os.path.exists(outputPath): 
		os.remove(outputPath)

	fileWrite = open(outputPath , "a" , encoding='utf_8_sig')
	
	WriteFileDescription(fileWrite , "ConfigManager.cpp" , "ConfigManager数据管理文件实现")
	fileWrite.write("#include \"" + "ConfigManager.h\"\n") 
	fileWrite.write("#include \"LogLib/inc/Log.h\"\n") 
	for sheet , item in g_xlsRecords.items():
		fileWrite.write("#include \"" + sheet + ".h\"\n") 
	fileWrite.write(oneTab + "\n\n") 

	fileWrite.write("namespace " + g_xlsNamespace + "\n") 
	fileWrite.write("{\n") 
	fileWrite.write(oneTab + "ConfigManager::ConfigManager()\n") 
	fileWrite.write(oneTab + "{\n") 
#	for sheet , item in g_xlsRecords.items():
#		if sheet in g_xlsDeleteRecord:
#			fileWrite.write("// " + twoTab + g_globaleNamePrefix + sheet + " = new " + g_xlsNamespace + "::" + sheet + ";\n") 
#		else:
#			fileWrite.write(twoTab + g_globaleNamePrefix + sheet + " = new " + g_xlsNamespace + "::" + sheet + ";\n") 
	fileWrite.write(oneTab + "}\n\n") 
	
	fileWrite.write(oneTab + "ConfigManager::~ConfigManager()\n") 
	fileWrite.write(oneTab + "{\n") 
#	for sheet , item in g_xlsRecords.items():
#		if sheet in g_xlsDeleteRecord:
#			pass
#		#	fileWrite.write("// " + twoTab + "SAFE_DELETE(" + g_xlsNamespace + "::" +  g_globaleNamePrefix + sheet + ");\n") 
#		else:
#		#	fileWrite.write(twoTab + "SAFE_DELETE(" + g_xlsNamespace + "::" +  g_globaleNamePrefix + sheet + ");\n") 
#			fileWrite.write(twoTab + g_xlsNamespace + "::" + sheet + "::GetInstance().Cleanup();\n") 
	fileWrite.write(oneTab + "}\n\n") 

	fileWrite.write(oneTab + "ConfigManager & ConfigManager::GetInstance()\n") 
	fileWrite.write(oneTab + "{\n") 
	fileWrite.write(twoTab + "static ConfigManager instance;\n") 
	fileWrite.write(twoTab + "return instance;\n") 
	fileWrite.write(oneTab + "}\n\n") 

	fileWrite.write(oneTab + "INT32 ConfigManager::Init(std::string  strCsvPath)\n") 
	fileWrite.write(oneTab + "{\n") 
	fileWrite.write(twoTab + "MsgAssert_ReF1(strCsvPath.length(), \"ConfigManager::Init error.\");\n\n") 
	fileWrite.write(twoTab + "if (strCsvPath[strCsvPath.length() - 1] != '/')\n") 
	fileWrite.write(twoTab + "{\n") 
	fileWrite.write(threeTab + "strCsvPath = strCsvPath + \"/\";\n") 
	fileWrite.write(twoTab + "}\n\n") 
	fileWrite.write(twoTab + "LoadFrom(strCsvPath);\n") 
	fileWrite.write(twoTab + "RepairLoad(strCsvPath);\n") 
	
	fileWrite.write(oneTab + "\n") 
	fileWrite.write(twoTab + "return 0;\n") 
	fileWrite.write(oneTab + "}\n\n") 

	fileWrite.write(oneTab + "INT32 ConfigManager::LoadFrom(std::string  strCsvPath)\n") 
	fileWrite.write(oneTab + "{\n") 
	fileWrite.write(twoTab + "MsgAssert_ReF1(strCsvPath.length(), \"ConfigManager::LoadFrom error.\");\n\n") 
	fileWrite.write(twoTab + "if (strCsvPath[strCsvPath.length() - 1] != '/')\n") 
	fileWrite.write(twoTab + "{\n") 
	fileWrite.write(threeTab + "strCsvPath = strCsvPath + \"/\";\n") 
	fileWrite.write(twoTab + "}\n\n") 
	for sheet , item in g_xlsRecords.items():
		if sheet in g_xlsDeleteRecord:
			fileWrite.write("// " + twoTab + "" + g_xlsNamespace + "::" + sheet + "::GetInstance().LoadFrom(strCsvPath);\n") 
		else:
			fileWrite.write(twoTab + "" + g_xlsNamespace + "::" + sheet + "::GetInstance().LoadFrom(strCsvPath);\n") 
	fileWrite.write(twoTab + "return 0;\n") 
	fileWrite.write(oneTab + "}\n\n") 
	
	
	fileWrite.write(oneTab + "INT32 ConfigManager::RepairLoad(std::string  strCsvPath)\n") 
	fileWrite.write(oneTab + "{\n") 
	fileWrite.write(twoTab + "MsgAssert_ReF1(strCsvPath.length(), \"ConfigManager::RepairLoad error.\");\n\n") 
	fileWrite.write(twoTab + "if (strCsvPath[strCsvPath.length() - 1] != '/')\n") 
	fileWrite.write(twoTab + "{\n") 
	fileWrite.write(threeTab + "strCsvPath = strCsvPath + \"/\";\n") 
	fileWrite.write(twoTab + "}\n\n") 
	for sheet , item in g_xlsRecords.items():
		if sheet in g_xlsDeleteRecord:
			fileWrite.write("// " + twoTab + "" + g_xlsNamespace + "::" + sheet + "::GetInstance().RepairLoad(strCsvPath);\n") 
		else:
			fileWrite.write(twoTab + "" + g_xlsNamespace + "::" + sheet + "::GetInstance().RepairLoad(strCsvPath);\n") 
	fileWrite.write(twoTab + "return 0;\n") 
	fileWrite.write(oneTab + "}\n\n") 
	
	ExportClassToLua(bServer , fileWrite)
	
	fileWrite.write(oneTab + "INT32 ConfigManager::Cleanup()\n") 
	fileWrite.write(oneTab + "{\n") 
#	for sheet , item in g_xlsRecords.items():
#		if sheet in g_xlsDeleteRecord:
#			pass
#		#	fileWrite.write("// " + twoTab + "SAFE_DELETE(" + g_xlsNamespace + "::" +  g_globaleNamePrefix + sheet + ");\n") 
#		else:
#		#	fileWrite.write(twoTab + "SAFE_DELETE(" + g_xlsNamespace + "::" +  g_globaleNamePrefix + sheet + ");\n") 
#			fileWrite.write(twoTab + g_xlsNamespace + "::" + sheet + "::GetInstance().Cleanup();\n") 
	fileWrite.write(twoTab + "return 0;\n") 
	fileWrite.write(oneTab + "}\n\n") 

	fileWrite.write("}\n\n") 
	
	fileWrite.close()

def ExportClassToLua(bServer , fileWrite): 
	if g_isExportLua:
		fileWrite.write(oneTab + "INT32 ConfigManager::ExportClassToLua(lua_State* L)\n") 
		fileWrite.write(oneTab + "{\n") 
		fileWrite.write(twoTab + "if (L == NULL)\n") 
		fileWrite.write(twoTab + "{\n") 
		fileWrite.write(threeTab + "return -1;\n") 
		fileWrite.write(twoTab + "}\n") 
		fileWrite.write(twoTab + "\n") 
		fileWrite.write(twoTab + "lua_tinker::namespace_add(L, \"" + g_xlsNamespace + "\");\n") 
		fileWrite.write(twoTab + "lua_tinker::namespace_add(L, \"" + "CUtil" + "\");\n") 
		fileWrite.write(twoTab + "lua_tinker::class_add<CUtil::CSVConfig>(L, \"CUtil::CSVConfig\", true);\n") 
		fileWrite.write(twoTab + "lua_tinker::class_add<CUtil::CSVConfig>(L, \"CUtil::CSVConfig\", true);\n") 
		fileWrite.write(twoTab + "lua_tinker::class_def<CUtil::CSVConfig>(L, \"IsLoaded\", &CUtil::CSVConfig::IsLoaded);\n") 
		fileWrite.write(twoTab + "\n") 
		for sheetIndex , sheet in g_xlsRecords.items():
			fileWrite.write(twoTab + "\n") 
			if sheetIndex in g_xlsDeleteRecord:
				fileWrite.write("//" + twoTab + "lua_tinker::class_add<" + g_xlsNamespace + "::" + sheetIndex + g_configSuffix + ">(L, \"" + g_xlsNamespace + "::" + sheetIndex + g_configSuffix + "\", true);\n") 
				fileWrite.write("//" + twoTab + "lua_tinker::class_add<" + g_xlsNamespace + "::" + sheetIndex + ">(L, \"" + g_xlsNamespace + "::" + sheetIndex + "\", true);\n") 
				fileWrite.write("//" + twoTab + "lua_tinker::scope_inner(L, \"" + g_xlsNamespace + "\" , \"" + sheetIndex + "\" , \"" + g_xlsNamespace + "::" + sheetIndex + "\");\n") 
				fileWrite.write("//" + twoTab + "lua_tinker::class_inh<" + g_xlsNamespace + "::" + sheetIndex + " , " + g_xlsNamespace + "::" + sheetIndex + g_configSuffix + ">(L);\n") 
				fileWrite.write("//" + twoTab + "lua_tinker::class_def_static<" + g_xlsNamespace + "::" + sheetIndex + ">(L, \"GetInstance\", & " + g_xlsNamespace + "::" + sheetIndex + "::GetInstance);\n") 
				fileWrite.write("//" + twoTab + "lua_tinker::class_def_static<" + g_xlsNamespace + "::" + sheetIndex + ">(L, \"GetInstancePtr\", & " + g_xlsNamespace + "::" + sheetIndex + "::GetInstancePtr);\n") 
				fileWrite.write("//" + twoTab + "lua_tinker::class_def<" + g_xlsNamespace + "::" + g_configPrefix + sheetIndex + ">(L, \"Get" + sheetIndex + "\", & " + g_xlsNamespace + "::" + g_configPrefix + sheetIndex + "::Get" + sheetIndex + ");\n") 
				fileWrite.write("//" + twoTab + "\n") 
				
				fileWrite.write("//" + twoTab + "lua_tinker::class_add<" + g_xlsNamespace + "::" + g_configPrefix + sheetIndex + ">(L, \"" + g_xlsNamespace + "::" + g_configPrefix + sheetIndex + "\", true);\n") 
				fileWrite.write("//" + twoTab + "lua_tinker::scope_inner(L, \"" + g_xlsNamespace + "\" , \"" + g_configPrefix + sheetIndex + "\" , \"" + g_xlsNamespace + "::" + g_configPrefix + sheetIndex + "\");\n") 
				#fileWrite.write(twoTab + "\n") 
			else:
				fileWrite.write(twoTab + "lua_tinker::class_add<" + g_xlsNamespace + "::" + sheetIndex + g_configSuffix + ">(L, \"" + g_xlsNamespace + "::" + sheetIndex + g_configSuffix + "\", true);\n") 
				fileWrite.write(twoTab + "lua_tinker::class_add<" + g_xlsNamespace + "::" + sheetIndex + ">(L, \"" + g_xlsNamespace + "::" + sheetIndex + "\", true);\n") 
				fileWrite.write(twoTab + "lua_tinker::scope_inner(L, \"" + g_xlsNamespace + "\" , \"" + sheetIndex + "\" , \"" + g_xlsNamespace + "::" + sheetIndex + "\");\n") 
				fileWrite.write(twoTab + "lua_tinker::class_inh<" + g_xlsNamespace + "::" + sheetIndex + " , " + g_xlsNamespace + "::" + sheetIndex + g_configSuffix + ">(L);\n") 
				fileWrite.write(twoTab + "lua_tinker::class_def_static<" + g_xlsNamespace + "::" + sheetIndex + ">(L, \"GetInstance\", & " + g_xlsNamespace + "::" + sheetIndex + "::GetInstance);\n") 
				fileWrite.write(twoTab + "lua_tinker::class_def_static<" + g_xlsNamespace + "::" + sheetIndex + ">(L, \"GetInstancePtr\", & " + g_xlsNamespace + "::" + sheetIndex + "::GetInstancePtr);\n") 
				fileWrite.write(twoTab + "lua_tinker::class_def<" + g_xlsNamespace + "::" + sheetIndex + ">(L, \"Get" + sheetIndex + "\", & " + g_xlsNamespace + "::" + sheetIndex + "::Get" + sheetIndex + " , std::string(\"\"));\n") 
				fileWrite.write(twoTab + "\n") 

				fileWrite.write(twoTab + "lua_tinker::class_add<" + g_xlsNamespace + "::" + g_configPrefix + sheetIndex + ">(L, \"" + g_xlsNamespace + "::" + g_configPrefix + sheetIndex + "\", true);\n") 
				fileWrite.write(twoTab + "lua_tinker::scope_inner(L, \"" + g_xlsNamespace + "\" , \"" + g_configPrefix + sheetIndex + "\" , \"" + g_xlsNamespace + "::" + g_configPrefix + sheetIndex + "\");\n") 
				#fileWrite.write(twoTab + "\n") 
				
				types = sheet[g_rowType]
				datas = sheet[g_rowName]
				css = sheet[g_rowCS]
				for index , item in enumerate(types):
					if not CheckCSType(css[index] , bServer):
						continue
					itemType = GetType(item)
					if itemType == g_structType or itemType == g_structArrayType:
						npos = datas[index].find('[')
						structName = datas[index][0 : npos]
						structDatas = datas[index][npos + 1 : len(datas[index]) - 1].split(',')
						childItems = item.split(',')

						fileWrite.write(twoTab + "{\n") 
						content =  g_xlsNamespace + "::" + g_configPrefix + sheetIndex + "::" + g_configPrefix + structName
						fileWrite.write(threeTab + "lua_tinker::class_add<" + content + ">(L, \"" + content + "\", true);\n") 
						#fileWrite.write(threeTab + "lua_tinker::class_inh<" + content + ", " + g_xlsNamespace + "::" + g_configPrefix + sheetIndex + ">(L);\n") 
						fileWrite.write(threeTab + "lua_tinker::scope_inner(L , \"" + g_xlsNamespace + "::" + g_configPrefix + sheetIndex  + "\",\"" + g_configPrefix + structName + "\" , \"" + content + "\");\n") 			

						for indexChild , childItem in enumerate(childItems):
							childItem = RemoveSpecialWord(childItem)
							fileWrite.write(threeTab + "lua_tinker::class_mem<" + content +">(L, \"" + structDatas[indexChild] + "\", &" + content + "::" + structDatas[indexChild] + ");\n") 
						fileWrite.write(twoTab + "}\n")

						if itemType == g_structArrayType:
							fileWrite.write(twoTab + "lua_tinker::class_mem<" + g_xlsNamespace + "::" + g_configPrefix + sheetIndex +">(L, \"vec" + structName+ "\", &" + g_xlsNamespace + "::" + g_configPrefix + sheetIndex + "::vec" + structName + ");\n") 
						if itemType == g_structType:
							fileWrite.write(twoTab + "lua_tinker::class_mem<" + g_xlsNamespace + "::" + g_configPrefix + sheetIndex +">(L, \"" + structName+ "\", &" + g_xlsNamespace + "::" + g_configPrefix + sheetIndex + "::" + structName + ");\n") 
					else:
						content =  g_xlsNamespace + "::" + g_configPrefix + sheetIndex
						fileWrite.write(twoTab + "lua_tinker::class_mem<" + content +">(L, \"" + datas[index] + "\", &" + content + "::" + datas[index] + ");\n") 
										
		fileWrite.write(twoTab + "return 0;\n") 
		fileWrite.write(oneTab + "}\n\n") 
	 
	for index , item in enumerate(types):
		#LogOutDebug("bServer:" , bServer , " data:" , datas[index] , "css:" , css[index])
		if not CheckCSType(css[index] , bServer):
			continue
		item_type = GetType(item)
		if item_type == g_configType:
			tmp = collections.OrderedDict()
			MakeDicKeyUpper(g_xlsRecords , tmp)
			parentName = GetDicKeyByUpper(g_xlsRecords , item)
			if parentName not in sameParaentName:
				sameParaentName[parentName] = parentName
				fileWrite.write("#include \"" + parentName + ".h\"\n") 
	fileWrite.write("\n") 
	
################################流程相关函数处理#####################################
def RemoveNewLine(str):
	pp = ""
	for tt in str.splitlines():
		tt = tt.strip().lstrip().rstrip()
		pp = pp+tt
	return pp

def RemovListNewLine(List): 
	for index , item2 in enumerate(List):
		List[index] = RemoveNewLine(item2)

def GetConfigType(item):
	tmp = collections.OrderedDict()
	MakeDicKeyUpper(g_xlsRecords , tmp)

	parentName = GetDicKeyByUpper(g_xlsRecords , item)
	parentType = g_xlsRecords[parentName][g_rowType][g_cellID] 
	valueType = g_configPrefix + parentName + " *"
	return g_configTypePrefix + GetType(parentType) + " , " + valueType + ">" , parentType , valueType

def GetType(item):
	tmp = collections.OrderedDict()
	MakeDicKeyUpper(g_xlsRecords , tmp)
	
	if item.upper() in tmp:
		return g_configType	
	elif  item.lower() == "int".lower() or\
		item.lower() == "int32".lower():
		return g_int32Type
	elif item.lower() == "int64".lower() or\
		item.lower() == "long".lower():
		return g_int64Type
	elif item.lower() == "double".lower() or\
		item.lower() == "float".lower():
		return g_doubleType
	elif item.lower() == "bool".lower():
		return g_boolType
	elif item.lower() == "string".lower() or\
		item.lower() == "std::string".lower():
		return g_stringType
	elif item.lower() == "condition".lower():
		return g_conditionType
	elif item.lower() == "date".lower() or\
		item.lower() == "Timer::date".lower():
		return g_dateType
	elif item.lower() == "int[]".lower() or\
		item.lower() == "[]int".lower() or\
		item.lower() == "[]int32".lower() or\
		item.lower() == "int32[]".lower():
		return g_int32ArrayType
	elif item.lower() == "int64[]".lower() or\
		item.lower() == "[]int64".lower() or\
		item.lower() == "[]long".lower() or\
		item.lower() == "long[]".lower():
		return g_int64ArrayType
	elif item.lower() == "double[]".lower() or\
		item.lower() == "[]double".lower() or\
		item.lower() == "[]float".lower() or\
		item.lower() == "float[]".lower():
		return g_doubleArrayType
	elif item.lower() == "date[]".lower() or\
		item.lower() == "[]date".lower()or\
		item.lower() == "Timer::date[]".lower()or\
		item.lower() == "[]Timer::date".lower():
		return g_dateArrayType
	elif item.lower() == "string[]".lower() or\
		item.lower() == "[]string".lower() or\
		item.lower() == "[]std::string".lower() or\
		item.lower() == "std::string[]".lower():
		return g_stringArrayType
	elif item.lower() == "bool[]".lower() or\
		item.lower() == "[]bool".lower():
		return g_boolArrayType
	elif item.lower() == "Condition".lower():
		return g_boolArrayType
	elif item.lower().find(',') >= 0 and \
		item.lower().find("<") == -1 and\
		item.lower().find(">") == -1 and\
		item.lower().find("[") == -1 and\
		item.lower().find("]") == -1:
		return g_structType
	elif item.lower().find(',') >= 0 and \
		item.lower()[0] == "[" and\
		item.lower()[len(item.lower()) - 1] == "]":
		return g_structArrayType
	elif item.lower().find('map') >= 0 and \
		item.lower().find("<") >= 0 and\
		item.lower()[len(item.lower()) - 1] == ">":
		mapType , keyType , valueType = GetMapType(item)
		#LogOutDebug("map" , item , "mapType:" , mapType , " key:" , keyType , "value:" , valueType)
		return mapType
	else:
		LogOutError("GetType error." , item)

	return "None"

def GetTypeFunc(item):
	tmp = collections.OrderedDict()
	MakeDicKeyUpper(g_xlsRecords , tmp)
	
	if item.upper() in tmp:
		return g_configFunc
	elif  item.lower() == "int".lower() or\
		item.lower() == "int32".lower():
		return g_int32Func
	elif item.lower() == "int64".lower() or\
		item.lower() == "long".lower():
		return g_int64Func
	elif item.lower() == "double".lower() or\
		item.lower() == "float".lower():
		return g_doubleFunc
	elif item.lower() == "bool".lower():
		return g_boolFunc
	elif item.lower() == "string".lower() or\
		item.lower() == "std::string".lower():
		return g_stringFunc
	elif item.lower() == "date".lower()or\
		item.lower() == "Timer::date".lower():
		return g_dateFunc
	elif item.lower() == "int[]".lower() or\
		item.lower() == "[]int".lower() or\
		item.lower() == "[]int32".lower() or\
		item.lower() == "int32[]".lower():
		return g_int32ArrayFunc
	elif item.lower() == "int64[]".lower() or\
		item.lower() == "[]int64".lower() or\
		item.lower() == "[]long".lower() or\
		item.lower() == "long[]".lower():
		return g_int64ArrayFunc
	elif item.lower() == "double[]".lower() or\
		item.lower() == "[]double".lower() or\
		item.lower() == "[]float".lower() or\
		item.lower() == "float[]".lower():
		return g_doubleArrayFunc
	elif item.lower() == "string[]".lower() or\
		item.lower() == "[]string".lower() or\
		item.lower() == "[]std::string".lower() or\
		item.lower() == "std::string[]".lower():
		return g_stringArrayFunc
	elif item.lower() == "bool[]".lower() or\
		item.lower() == "[]bool".lower():
		return g_boolArrayFunc
	elif item.lower() == "Date[]".lower() or\
		item.lower() == "[]Date".lower()or\
		item.lower() == "Timer::date[]".lower()or\
		item.lower() == "[]Timer::date".lower():
		return g_dateArrayFunc
	elif item.lower().find(',') >= 0 and \
		item.lower().find("<") == -1 and\
		item.lower().find(">") == -1 and\
		item.lower().find("[") == -1 and\
		item.lower().find("]") == -1:
		return g_structFunc
	elif item.lower().index(",") >= 0 and \
		item.lower()[0] == "[" and\
		item.lower()[len(item.lower()) - 1] == "]":
		return g_structArrayFunc
	elif item.lower().find('map') >= 0 and \
		item.lower().find("<") >= 0 and\
		item.lower()[len(item.lower()) - 1] == ">":
		return g_mapFunc
	else:
		LogOutError("GetTypeFunc error." , item)

	return None
	
def GetMapType(item , bReadConfigType = True):
	if item.lower().find('map') >= 0 and \
		item.lower().find("<") and\
		item.lower()[len(item.lower()) - 1] == ">":
		item = RemoveSpecialWord(item)
		childItems = item.split('<')[1].rstrip('>').split(',')
		keyType = GetType(RemoveSpecialWord(childItems[0]))
		keyName = keyType
		if keyType == g_configType:
			keyName = GetDicKeyByUpper(g_xlsRecords , childItems[0])
			parentType = g_xlsRecords[keyName][g_rowType][g_cellID]
			keyType = GetType(parentType)

		valueType = GetType(RemoveSpecialWord(childItems[1]))
		valueName = valueType
		if valueType == g_configType:
			valueName = GetDicKeyByUpper(g_xlsRecords , childItems[1])
			parentType = g_xlsRecords[valueName][g_rowType][g_cellID]			
			valueType = GetType(parentType)

		if bReadConfigType:
			return "std::map<"+ keyType + " , " + valueType + ">" , keyType , valueType
		else:
			return "std::map<"+ keyType + " , " + valueType + ">" , keyName , valueName

def GetTypeTab(item):
	tmp = collections.OrderedDict()
	MakeDicKeyUpper(g_xlsRecords , tmp)
	
	if item.upper() in tmp:
		return oneTab	
	if  item.lower() == "int".lower() or\
		item.lower() == "int32".lower():
		return sixTab
	elif item.lower() == "bool".lower():
		return sixTab
	elif item.lower() == "int64".lower() or\
		item.lower() == "long".lower():
		return sixTab
	elif item.lower() == "double".lower() or\
		item.lower() == "float".lower():
		return sixTab
	elif item.lower() == "Date".lower()or\
		item.lower() == "Timer::date".lower():
		return fiveTab
	elif item.lower() == "string".lower() or\
		item.lower() == "std::string".lower():
		return fiveTab
	elif item.lower() == "int[]".lower() or\
		item.lower() == "[]int".lower() or\
		item.lower() == "[]int32".lower() or\
		item.lower() == "int32[]".lower():
		return threeTab
	elif item.lower() == "bool[]".lower() or\
		item.lower() == "[]bool".lower():
		return threeTab
	elif item.lower() == "int64[]".lower() or\
		item.lower() == "[]int64".lower() or\
		item.lower() == "[]long".lower() or\
		item.lower() == "long[]".lower():
		return threeTab
	elif item.lower() == "double[]".lower() or\
		item.lower() == "[]double".lower() or\
		item.lower() == "[]float".lower() or\
		item.lower() == "float[]".lower():
		return threeTab
	elif item.lower() == "string[]".lower() or\
		item.lower() == "[]string".lower() or\
		item.lower() == "[]std::string".lower() or\
		item.lower() == "std::string[]".lower():
		return oneTab
	elif item.lower() == "Date[]".lower() or\
		item.lower() == "[]Date".lower()or\
		item.lower() == "Timer::date[]".lower()or\
		item.lower() == "[]Timer::date".lower():
		return twoTab
	elif item.lower().find(',') >= 0 and \
		item.lower().find("<") == -1 and\
		item.lower().find(">") == -1 and\
		item.lower().find("[") == -1 and\
		item.lower().find("]") == -1:
		return twoTab
	elif item.lower().index(",") >= 0 and \
		item.lower()[0] == "[" and\
		item.lower()[len(item.lower()) - 1] == "]":
		return oneTab
	if item.lower().find('map') >= 0 and \
		item.lower().find("<") >= 0 and\
		item.lower()[len(item.lower()) - 1] == ">":
		return oneTab
	else:
		return twoTab

	return oneTab

def GetDateType(date):
	dateItems = date.split('-')
	if len(dateItems) == 4:
		return "Timer::DATE_TYPE_YEAR"
	elif len(dateItems) == 3:
		return "Timer::DATE_TYPE_MON"
	elif len(dateItems) == 2:
		return "Timer::DATE_TYPE_DAY"
	elif len(dateItems) == 1:
		timeItems = date.split(':')
		if len(timeItems) == 3:
			return "Timer::DATE_TYPE_HOUR"
		elif len(timeItems) == 2:
			return "Timer::DATE_TYPE_MIN"
		elif len(timeItems) == 1:
			return "Timer::DATE_TYPE_SEC"
	
	LogOutError("GetDateType error ." , date)
	return "Timer::DATE_TYPE_INVALID"
	
def RemoveSpecialWord(item):
	item = item.replace(' ','').replace('	','').replace('[','').replace(']','').strip().rstrip()
	if len(item) > 0 and item[len(item) - 1] == ',':
		#LogOutDebug("item:" , item)
		item = item[0:len(item) - 1]
		#LogOutDebug("end item:" , item)
	if len(item) > 0 and item[0] == ',':
		#LogOutDebug("item:" , item)
		item = item[1:]
		#LogOutDebug("end item:" , item)
	return item

def CheckDataArray(colItem):
	pp = ""
	childItems = colItem.split(',')
	for childIndex , childItem in enumerate(childItems):							
		childItem = RemoveSpecialWord(childItem)
		if len(childItem) > 0:
			pp = pp + childItem
			if len(childItems) - 1 != childIndex:
				pp = pp + ","

	#LogOutInfo("pp " , pp) 
	return pp

def CheckDataType(item_type , sheet , row , col , colItem , childIndex):	
	if item_type == g_boolType:
		item = RemoveSpecialWord(colItem)
		if item.lower() == "true".lower() or\
			item.lower() == "false".lower() or\
			item.lower() == "0".lower() or\
			item.lower() == "1".lower():	
			pass
		else:
			LogOutError("sheet=" , sheet , " :row=" , row , " :col=" , col , " item=" , item  , " bool type error.")							

		if row != -1:
			g_xlsRecords[sheet][row][col] = item
		else:
			return item
	elif item_type == g_int32Type:
		item = RemoveSpecialWord(colItem)
		if not item.lower().isdigit():	
			LogOutError("sheet=" , sheet , " :row=" , row , " :col=" , col , " item=" , item  , " int32 type error.")							

		if row != -1:
			g_xlsRecords[sheet][row][col] = item
		else:
			return item
	elif item_type == g_int64Type:
		item = RemoveSpecialWord(colItem)
		if not item.lower().isdigit():	
			LogOutError("sheet=" , sheet , " :row=" , row , " :col=" , col , " item=" , item , " int64 type error.")							

		if row != -1:
			g_xlsRecords[sheet][row][col] = item
		else:
			return item
	elif item_type == g_doubleType:						
		item = RemoveSpecialWord(colItem)
		#if !item.lower().isdigit():	
		#	LogOutError("sheet=" , sheet , " :row=" , row , " :col=" , col , " item=" , " double type error.")							

		if row != -1:
			g_xlsRecords[sheet][row][col] = item
		else:
			return item
	elif item_type == g_stringType:
		return colItem
	elif item_type == g_conditionType:
		item = RemoveSpecialWord(colItem)
		if row >= g_conditionDataStart:
			tagitem = item.lower()[0:item.lower().find(':')]
			itemContent = item.lower()[item.lower().find(':') + 1:]
			#LogOutDebug("condition item:" , item , " tagitem=" , tagitem , " itemContent " , itemContent)
			if tagitem.lower() == g_conditionTag.lower() or tagitem.lower() == g_actionTag.lower():	
				pass
			else:
				LogOutError("sheet=" , sheet , " :row=" , row , " :col=" , col , " item=" , item , " condition tag error.")						

			if row != -1:
				g_xlsConditionRecords[sheet][col][row] = item
			else:
				return item
		return item

	elif item_type == g_dateType:
		item = RemoveSpecialWord(colItem)
		if row != -1:
			g_xlsRecords[sheet][row][col] = item
		return item
	elif item_type == g_dateArrayType:
		g_xlsRecords[sheet][row][col] = CheckDataArray(colItem)
	elif item_type == g_int32ArrayType:
		g_xlsRecords[sheet][row][col] = CheckDataArray(colItem)
	elif item_type == g_int64ArrayType:
		g_xlsRecords[sheet][row][col] = CheckDataArray(colItem)
	elif item_type == g_doubleArrayType:
		g_xlsRecords[sheet][row][col] = CheckDataArray(colItem)
	elif item_type == g_dateArrayType:
		g_xlsRecords[sheet][row][col] = CheckDataArray(colItem)
	elif item_type == g_stringArrayType:
		return colItem
	elif item_type == g_configType:
		name = g_xlsRecords[sheet][g_rowType][col]

		if GetType(name) == g_structType or GetType(name) == g_structArrayType:
			name = RemoveSpecialWord(name)			
			itemContent = name.split(',')		
			name = itemContent[childIndex]

		parentName = GetDicKeyByUpper(g_xlsRecords , name)
		#LogOutDebug("name:" , name , "sheet:" , sheet , "parentName:" , parentName)
		parentType = g_xlsRecords[parentName][g_rowType][g_cellID]
		if not CheckConfigData(parentName , colItem):
			LogOutError(colItem , " not in " , parentName , " id col." , "config row=" , row , " col=" , col)
		return CheckDataType(parentType , sheet , row , col , colItem , childIndex)
	elif item_type == g_conditionType:
		item = RemoveSpecialWord(colItem)
		if item.index("(") >= 0 or item.index(")") >= 0 or item.index("|") >= 0:	
			LogOutError("sheet=" , sheet , " :row=" , row , " :col=" , col , " item=" , item  , " Condition  type error.")				

		if row != -1:
			g_xlsRecords[sheet][row][col] = item
		else:
			return item
			
		return item
	elif item_type == g_structType:
#		if len(colItem) > 0 and colItem[len(colItem) - 1] == ',':
#			colItem = colItem[0:len(colItem) - 1]
#		if len(colItem) > 0 and colItem[0] == ',':
#			colItem = colItem[1:]
		childItems = colItem.split(',')		
		itemContent = ""
		for childIndex , childItem in enumerate(childItems):
			if len(childItem) == 0:
				continue
			childItem = RemoveSpecialWord(childItem)
			#LogOutDebug(" childItem" , childItem)
			childType = GetTypeByIndex(g_xlsRecords[sheet][g_rowType][col] , childIndex)
			#LogOutDebug("childType" , childType , " childItem" , childItem)
			childItem = CheckDataType(childType , sheet , -1 , col , childItem , childIndex)
			itemContent = itemContent + childItem
			if len(childItems) - 1 != childIndex:
				itemContent = itemContent + ","

		#LogOutDebug("itemContent:" , itemContent)
		if row != -1:
			g_xlsRecords[sheet][row][col] = itemContent
		else:
			item = "["
			item = item + itemContent
			item = item + "]"
			return item

	elif item_type == g_structArrayType:
		if row != -1:
			childItems = colItem.split(']')		
			itemContent = ""
			for childIndex , childItem in enumerate(childItems):
				if len(childItem) > 0:
					#LogOutDebug("itemContent:" , itemContent , "childIndex" , childIndex , "size=" , len(childItems) , "childItems" , childItems , "colItem:" , colItem)
					childItem = childItem.replace('[' , '').replace(']' , '').strip().rstrip().lstrip()
					childItem = CheckDataType(g_structType , sheet , -1 , col , childItem , -1)
					itemContent = itemContent + childItem

			g_xlsRecords[sheet][row][col] = itemContent
		else:
			return colItem
	else:
		#LogOutDebug("item_type" , item_type)
		if item_type.find(g_mapType) >= 0:  #Map类型
			childItems = colItem.split('=')		
			itemContent = ""
			for childIndex , childItem in enumerate(childItems):
				if childIndex == 0:
					itemType = RemoveSpecialWord(childItem)
					itemContent = itemContent + itemType				
				else:
					itemContent = itemContent + "="
					itemContent = itemContent + childItem.strip()
			#LogOutDebug("itemContent" , itemContent)
			if row != -1:
				g_xlsRecords[sheet][row][col] = itemContent
			return itemContent
		else:
			return colItem

def CheckConfigData(parentName , colItem):
	for row , rowItem in g_xlsRecords[parentName].items():		#读取每一行
		if rowItem[g_cellID] == colItem:
			return True
	return False
	
def GetTypeByIndex(types , index):
	typeItems = types.split(',')
	for childIndex , childItem in enumerate(typeItems):
		childItem = RemoveSpecialWord(childItem)
		if index == childIndex:
			return GetType(childItem)

	LogOutError("GetTypeByIndex error.types=" , types , " index=" ,index)

def MakeTitle(types , datas ):
	Str = "// attention dont't change this line:"		
	for index , item in enumerate(types):
		item_type = GetType(item)
		if item_type == g_structType:
			npos = datas[index].find('[')
			structName = datas[index][0 : npos]
			structDatas = datas[index][npos + 1 : len(datas[index]) - 1].split(',')
			childItems = item.split(',')
			Str = Str + structName
			if len(childItems) != len(structDatas):
				LogOutError("parase struct error.invalid size.index=" , index , " . name=" , structName , ":childItems " , childItems , "size=" , len(childItems) , ":structDatas " , structDatas , "size=" , len(structDatas))
			for indexChild , childItem in enumerate(childItems):
				childType = GetType(childItem)
				if childType == g_configType:
					parentName = GetDicKeyByUpper(g_xlsRecords , childItem)
					childType , keyType , valueType  = GetConfigType(parentName)

				Str = Str + childType + " " + structDatas[indexChild]  + ";"
		elif item_type == g_structArrayType:
			npos = datas[index].find('[')
			structName = datas[index][0 : npos]
			structDatas = datas[index][npos + 1 : len(datas[index]) - 1].split(',')
			
			Str = Str + structName
			item = item.replace('[' , '').replace(']' , '')
			childItems = item.split(',')
			if len(childItems) != len(structDatas):
				LogOutError("parase struct error.invalid size.index=" , index , " . name=" , structName , ":childItems " , childItems , "size=" , len(childItems) , ":structDatas " , structDatas , "size=" , len(structDatas))
			for indexChild , childItem in enumerate(childItems):
				childType = GetType(childItem)
				if childType == g_configType:
					parentName = GetDicKeyByUpper(g_xlsRecords , childItem)
					childType , keyType , valueType  = GetConfigType(parentName)
				Str = Str + childType + " " + structDatas[indexChild]  + ";"		
		else:
			if item_type == g_configType:
				tmp = collections.OrderedDict()
				MakeDicKeyUpper(g_xlsRecords , tmp)
				item_type = GetDicKeyByUpper(g_xlsRecords , item)		
				#item_type , keyType , valueType = GetConfigType(item)
			Str = Str + item_type + " " + datas[index]   + ";"
	Str += "\n"
	return Str
	
def CheckNeedDelete(outputFile , types , datas):
	if os.path.exists(outputFile): 
		strNew = MakeTitle(types , datas)
		
		fileRead = open(outputFile , "r" , encoding = "utf_8_sig")
		for line in fileRead: 
			strOld = line
			break

		#LogOutDebug("\nold comments:" , strOld , "new comments:" , strNew , "\n")
		if strNew != strOld:
			LogOutInfo(outputFile , " not match.\n")
			LogOutInfo("old comments:" , strOld)
			LogOutInfo("new comments:" , strNew)		
		fileRead.close()
		return True
	else:
		return False

def GetCPPFilePath(bServer): 
	if bServer:
		return g_xlsExportServerCPPPath
	else:
		return g_xlsExportClientCPPPath
	
def IsServerCSType(data): 
	if data.lower().find(g_serverType) < 0 and data.lower().find(g_clientType) < 0:
		LogOutError("CheckCSType error." , data)
		
	if data.lower().find(g_serverType) >= 0:
		return True
	else:
		return False
		
def CheckCSType(data , bServer = True , bStrict = False): 
	if data.lower().find(g_serverType) < 0 and data.lower().find(g_clientType) < 0:
		LogOutError("CheckCSType error." , data)

	if bServer:
		if bStrict or g_bCheckCSTypeStrict:
			if data.lower().find(g_serverType) >= 0:
				return True
			else:
				return False
		else:
			return True # 服务器都需要
	
	else:
		if data.lower().find(g_clientType) >= 0:
			return True
		else:
			return False

def MakeDicKeyUpper(value , tmp):
	for index , item in value.items():
		tmp[index.upper()] = item
		
def GetDicKeyByUpper(value , tmp):
	for indexValue , item in value.items():
		if indexValue.upper() == tmp.upper():
			return indexValue
		
################################流程无关函数处理#####################################
def Usage():
    print('GenerateCSV.py usage:')
    print('-h,--help: print help message.')
    print('-v, --version: print script version')
    print('-o, --output: input an csv output verb')
    print('-c, --output: input an cpp output verb')
    print('--foo: Test option ')
    print('--fre: another test option')

def Version():
	print('GenerateCSV.py 1.0.0.0.1')

def LogOutDebug(*string):
	cp = GetColor("debug")
	longStr = cp + "[ DEBUG ]"
	for item in range(len(string)):  
		longStr += str(string[item])

	print(longStr)
	GetColor("reset")

def LogOutInfo(*string):
	cp = GetColor("info")
	longStr = cp + "[ INFO ] "
	for item in range(len(string)):  
		longStr += str(string[item])
	
	print(longStr)
	GetColor("reset")
	
def LogOutError(*string):
	cp = GetColor("error")
	longStr = cp + "[ ERR ] "
	for item in range(len(string)):  
		longStr += str(string[item])
	
	print(longStr)
	GetColor("reset")
	traceback.print_stack() 
	sys.exit()
	
def InitColor():
	if g_platform == "LINUX":
		cp = '\033['
		g_yellow = cp + '33m'
		g_yellow_h = cp + '1;33m'
		g_green = cp + '32m'
		g_green_h = cp + '1;32m'
		g_red = cp + '31m'
		g_cyan = cp + '36m'
		g_original = cp + '0m'
	else:
		g_yellow = ""
		g_yellow_h = ""
		g_green = ""
		g_green_h = ""
		g_red = ""
		g_original = ""
		g_cyan = ""

def GetColor(colorType):
	stdOutHandle = ctypes.windll.kernel32.GetStdHandle(g_stdOutputHandle)
	if colorType == "error":
		if g_platform == "LINUX":
			return g_red
		else:
			ctypes.windll.kernel32.SetConsoleTextAttribute(stdOutHandle , FOREGROUND_RED | FOREGROUND_INTENSTITY)
			return ""
	elif colorType == "info":
		if g_platform == "LINUX":
			return g_green
		else:
			ctypes.windll.kernel32.SetConsoleTextAttribute(stdOutHandle , FOREGROUND_GREEN | FOREGROUND_INTENSTITY)
			return ""
	elif colorType == "debug" or colorType == "reset":
		if g_platform == "LINUX":
			return g_original
		else:
			ctypes.windll.kernel32.SetConsoleTextAttribute(stdOutHandle , FOREGROUND_BLUE | FOREGROUND_RED | FOREGROUND_GREEN)
			return ""
	elif colorType == "warning" :
		if g_platform == "LINUX":
			return g_yellow
		else:
			ctypes.windll.kernel32.SetConsoleTextAttribute(stdOutHandle , FOREGROUND_BLUE | FOREGROUND_INTENSTITY)
			return ""
				
def WriteFileDescription(fileWrite , sfile , desc):
	fileWrite.write("/************************************" + "\n")
	fileWrite.write("FileName	:	" + sfile + "\n")
	fileWrite.write("Author		:	generate by tools" + "\n")
	fileWrite.write("HostName	:	" + socket.gethostname() + "\n")
	fileWrite.write("IP			:	" + socket.gethostbyname(socket.gethostname()) + "\n")
	fileWrite.write("Version		:	0.0.1" + "\n")
	fileWrite.write("Date		:	" + time.strftime('%Y-%m-%d %H:%M:%S') + "\n")
	fileWrite.write("Description	:	" + desc + "\n")
	fileWrite.write("************************************/" + "\n")

def Search(base , suffix , fileresult):
	pattern = suffix
	cur_list = os.listdir(base)
	for item in cur_list:
		full_path = os.path.join(base, item)
		if os.path.isdir(full_path):            
			Search(full_path , suffix, fileresult)
		if os.path:
			if full_path.endswith(pattern):
				CheckIsInList(fileresult , full_path)
				fileresult.append(full_path)
	return fileresult
	
def CheckIsInList(files , cur_file):
	filename = os.path.basename(cur_file) #获取文件名
	for sfile in files:
		tmp = os.path.basename(sfile) #获取文件名
		if tmp == filename:
			LogOutError("the same file. first=" , cur_file , " :second=" , sfile)


def DeleteExportPathFiles():
	files = []
	Search(g_xlsExportCSVPath ,  g_csvFileSuffix , files)
	for sfile in files:
		if os.path.exists(sfile): 
			os.remove(sfile)

def CreateExportPathFiles(): 
	if False == os.path.exists(g_xlsExportCSVPath):
		os.makedirs(g_xlsExportCSVPath)
		LogOutInfo("create dir: " , g_xlsExportCSVPath)
		
	if False == os.path.exists(g_xlsExportServerCPPPath):
		os.makedirs(g_xlsExportServerCPPPath)
		LogOutInfo("create dir: " , g_xlsExportServerCPPPath)

	if False == os.path.exists(g_xlsExportClientCPPPath):
		os.makedirs(g_xlsExportClientCPPPath)
		LogOutInfo("create dir: " , g_xlsExportClientCPPPath)

################################main函数处理#####################################
def handleArgs(argv): 
	global g_xlsImportPath 
	global g_xlsExportCSVPath
	global g_xlsExportServerCPPPath
	global g_xlsExportClientCPPPath
	
	try:
		opts, args = getopt.getopt(argv[1:], 'hvi:o:c:s:', ['import='])
	except: 
		Usage()
		sys.exit(2) 
	if len(argv) < 2: 
		Usage()
		return  
	for o, a in opts:
		if o in ('-h', '--help'):
			Usage()
			sys.exit(1)
		elif o in ('-v', '--version'):
			Version()
			sys.exit(0) 
		elif o in ('-i','--import',):
			g_xlsImportPath = a
		elif o in ('-o','--export',):
			g_xlsExportCSVPath = a
		elif o in ('-s','--servercpp',):
			g_xlsExportServerCPPPath = a
		elif o in ('-c','--clientcpp',):
			g_xlsExportClientCPPPath = a
		elif o in ('--fre',):
			Fre=a
		else:
			print('unhandled option')
			sys.exit(3) 
	if g_xlsImportPath == "":
		LogOutError("no xls g_xlsImportPath")
	elif g_xlsExportCSVPath == "":
		LogOutError("no xls g_xlsExportCSVPath")
			
def main(argv):
	global g_xlsImportPath 
	global g_xlsExportCSVPath
	global g_xlsExportServerCPPPath
	global g_xlsExportClientCPPPath
	
	InitColor()
	handleArgs(argv)
	g_xlsImportPath = "./xls_config"
	g_xlsExportCSVPath = "../../../bin/vs14.0/x64/Lib_Debug_x64/csv_config"
	g_xlsExportServerCPPPath = "../../../vsproject/TestLibCore/CSVConfigs"
	g_xlsExportClientCPPPath = "../../../vsproject/TestLibCore/CSVClientConfigs"
	LogOutInfo("generate csv.path:" + g_xlsImportPath + " csv:" + g_xlsExportCSVPath + " server_cpp:" + g_xlsExportServerCPPPath + " client_cpp:" + g_xlsExportClientCPPPath) 
	start()
	LogOutInfo("complete generate csv.") 
	
if __name__ == '__main__': 
	main(sys.argv)


